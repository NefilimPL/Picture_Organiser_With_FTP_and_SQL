# --- Konfigurowalne ustawienia ---
# Opcjonalna lokalizacja bazowa. Wstaw własną ścieżkę,
# aby wymusić zapis plików w innym katalogu niż domyślny "Pictures".
BASE_DIR_OVERRIDE = r""

# Klucz używany do prostego szyfrowania danych konfiguracyjnych.
APP_SECRET = "secret_v1"

# Domyślny port serwera FTP.
PORT = 21

# Zapytanie SQL aktualizujące ścieżkę do obrazu.
SQL_UPDATE_TEMPLATE = (
    
)

# Domyślne dane konfiguracyjne wykorzystywane przy pierwszym uruchomieniu.
DEFAULT_CONFIG = {
    "ftp": {
        "host": r"",
        "port": PORT,
        "user": r"",
        "pass": r"",
        "path": r"/PHOTOS/",
    },
    "sql": {
        "server": r"",
        "database": r"",
        "user": r"",
        "pass": r"",
    },
    "mysql": {
        "server": r"",
        "database": r"",
        "user": r"",
        "pass": r"",
    },
    "db_type": r"mysql",
    "sql_query": SQL_UPDATE_TEMPLATE,
    "enable_ftp_update": True,
    "enable_sql_update": True,
    "loc_path": r"",
}
# --- Koniec konfiguracji ---

# Komunikaty dla użytkownika
PROCESSING_MSG = "Trwa przetwarzanie. Poczekaj na zakończenie bieżącej operacji."
OPERATION_TITLE = "Operacja w toku"

# Podstawowe ustawienia
A_ = "1.0"
Az = "normal"
Ay = False
Ax = range
Al = True
Ak = "disabled"
Aj = getattr
AQ = None

# Komunikaty o błędach
NETWORK_ERROR_MSG = "Błąd sieciowy lub brak internetu"
PATH_NOT_FOUND_MSG = "Nie znaleziono ścieżki na serwerze"
NO_SUCH_FILE_MSG = "No such file"
LOGIN_DATA_ERROR_MSG = "Błędne dane logowania"
LOGIN_INCORRECT_MSG = "Login incorrect"
NO_DATA_MSG = "Brak danych"
MISSING_FIELDS_MSG = "Uzupełnij wszystkie wymagane pola przed dodaniem pliku."
INCOMPLETE_DATA_MSG = "Niekompletne dane"

# Oznaczenia interfejsu
CANCEL_LABEL = "Anuluj"
SETTINGS_LABEL = "Ustawienia"
EDIT_LISTS_LABEL = "Edytuj listy"
LIGHT_GREEN = "lightgreen"
OPEN_FURNITURE = "open_furniture"
NON_PIC = "non_pic"
ELEMENT_PIC = "element_pic"

# Klasy wyjątków
TIMEOUT_ERROR = TimeoutError
CONNECTION_REFUSED_ERROR = ConnectionRefusedError
As = "550"
Am = "left"
An = "vertical"
At = "PNG"
Ao = "Plik zablokowany"
Ap = "przez inny proces"
Aq = isinstance
Au = OSError
AR = "blue"
AS = "frame"
Aa = "prefix"
Ab = "red"
AT = "white"
NO_FILE_FALLBACK = "Brak pliku"
AV = "right"
Ac = "Błąd zapisu"
AW = "KOLOR3"
AX = "KOLOR2"
AY = "KOLOR1"
AZ = "MODEL"
Ad = "TYP"
Ae = "NAZWA"
AI = ", "
AJ = "Brak na liście"
AK = "Błąd"
A6 = "%Y-%m-%d %H:%M:%S"
A7 = "x_lbl"
A8 = "#aaa"
A4 = "green"
A2 = "<<ComboboxSelected>>"
y = "img_lbl"
z = "both"
A0 = enumerate
x = open
ft = "enable_ftp_update"
u = "enable_sql_update"
v = "host"
w = "sql_query"
p = "db_type"
q = "BRAK-EAN"
r = "port"
s = "MODELE"
t = "TYPY"
m = "path"
k = "utf-8"
n = "NAZWY"
j = "TCombobox"
f = "filepath"
B0 = "mark"
g = "-"
a = "_"
h = Ay
b = "database"
c = "server"
d = "DODATKI"
Z = "Existing.TCombobox"
Y = "KOLORY"
W = "ENTRIES"
R = "e"
T = "w"
S = "values"
V = Ak
Q = len
P = "sql"
X = Az
M = "pass"
N = "user"
L = "NO-LED"
K = "mysql"
J = Al
I = AQ
H = "ftp"
E = Exception
G = str
B = ""
import sys, os as A, subprocess as BH, shutil as Af, getpass, platform as BR, locale as BO
from datetime import datetime as A9
import time as Ag, tempfile, uuid
from tkinter import scrolledtext as BS
import threading, urllib.request as BN, urllib.parse as BP

AO = getpass.getuser()
AF = BR.node()
OLD_HOST_KEY = (AF or B) + "secret_OLD"

def ensure_package(pkg_name, import_name=I):
    A = pkg_name
    try:
        __import__(import_name or A)
    except ImportError:
        BH.check_call([sys.executable, "-m", "pip", "install", A])


ensure_package("tkinterdnd2")
ensure_package("Pillow", "PIL")
ensure_package("openpyxl")
ensure_package("pyodbc")
ensure_package("mysql-connector-python", "mysql.connector")
import tkinter as F
from tkinter import ttk as C, filedialog as BT, messagebox as O, simpledialog as BI
from tkinterdnd2 import TkinterDnD as BU, DND_ALL, DND_FILES as BJ
from PIL import Image as AA, ImageTk
from openpyxl import Workbook as BV, load_workbook as Ah
import ftplib as AB, socket as BK, pyodbc, mysql.connector, ctypes, json as Ar, base64 as BL

AC = BASE_DIR_OVERRIDE or A.path.join(A.path.expanduser("~"), "Pictures")
l = A.path.join(AC, "_ZDJECIA PRZEROBIONE_")
o = A.path.join(AC, "lists.xlsx")
AD = A.path.join(AC, "config.json")
AM = A.path.join(AC, "error_log.txt")
BM = A.path.join(AC, "changes_log.txt")
AN = A.path.join(AC, "temp_backup")
LC_DEFAULT = A.path.join(A.path.expanduser("~"), "Gui zdjęcia config", "Localization")
LC = LC_DEFAULT
# Domyślne odnośniki do plików lokalizacyjnych wykorzystywane,
# gdy w konfiguracji nie zdefiniowano własnej listy.
LOC_URLS_DEFAULT = [
    "https://github.com/NefilimPL/Picture_Organiser_With_FTP_and_SQL/tree/main/Localization"
    
]
LOC_URLS = LOC_URLS_DEFAULT[:]
DEFAULT_CONFIG["loc_path"] = LC_DEFAULT
DEFAULT_CONFIG["loc_urls"] = LOC_URLS_DEFAULT
AE = {n: n, t: t, s: s, Y: Y, d: d, W: W}
BW = [
    "ODBC Driver 18 for SQL Server",
    "ODBC Driver 17 for SQL Server",
    "SQL Server Native Client 11.0",
    "SQL Server",
]
Ai = OLD_HOST_KEY
AG = AQ


def set_app(app):
    global AG
    AG = app


def _xor_enc(s, key):
    if s is I or s == B:
        return B
    raw = B.join(chr(ord(ch) ^ ord(key[i % Q(key)])) for (i, ch) in A0(s))
    return BL.b64encode(raw.encode(k)).decode(k)


def _xor_dec(s, key):
    if s is I or s == B:
        return B
    try:
        raw = BL.b64decode(s.encode(k)).decode(k)
    except E:
        return s
    return B.join(chr(ord(ch) ^ ord(key[i % Q(key)])) for (i, ch) in A0(raw))


def encrypt(data):
    return _xor_enc(data, APP_SECRET)


def decrypt(enc_data):
    v = _xor_dec(enc_data, APP_SECRET)
    if not v or any(ord(ch) < 9 for ch in v):
        v2 = _xor_dec(enc_data, OLD_HOST_KEY)
        if v2 and all(ord(ch) >= 9 for ch in v2):
            return v2
    return v


def load_config():
    global AD
    B = Ar.loads(Ar.dumps(DEFAULT_CONFIG))
    config_path = AD
    if not A.path.exists(config_path):
        if not BASE_DIR_OVERRIDE:
            C_ = BT.askdirectory(title="Wskaż folder z plikiem konfiguracyjnym")
            if C_:
                config_path = A.path.join(C_, "config.json")
        if not A.path.exists(config_path):
            I_ = {
                H: {
                    v: B[H][v],
                    r: B[H][r],
                    N: encrypt(B[H][N]),
                    M: encrypt(B[H][M]),
                    m: B[H][m],
                },
                P: {c: B[P][c], b: B[P][b], N: encrypt(B[P][N]), M: encrypt(B[P][M])},
                K: {c: B[K][c], b: B[K][b], N: encrypt(B[K][N]), M: encrypt(B[K][M])},
                p: B[p],
                w: B[w],
                ft: B[ft],
                u: B[u],
                "loc_path": B.get("loc_path", LC_DEFAULT),
                "loc_urls": B.get("loc_urls", LOC_URLS_DEFAULT),
            }
            try:
                A.makedirs(A.path.dirname(config_path), exist_ok=J)
                with x(config_path, T, encoding=k) as D_:
                    Ar.dump(I_, D_, indent=4)
            except E as F_:
                try:
                    with x(AM, "a", encoding=k) as G_:
                        G_.write(
                            f"[{A9.now().strftime(A6)}] [USER: {AO}] [PC: {AF}] ERROR: Failed to create config.json: {F_}\n"
                        )
                except:
                    pass
    AD = config_path
    try:
        with x(AD, "r", encoding=k) as D_:
            C = Ar.load(D_)
        B[H][v] = C.get(H, {}).get(v, B[H][v])
        B[H][r] = C.get(H, {}).get(r, B[H][r])
        B[H][N] = decrypt(C.get(H, {}).get(N, encrypt(B[H][N])))
        B[H][M] = decrypt(C.get(H, {}).get(M, encrypt(B[H][M])))
        B[H][m] = C.get(H, {}).get(m, B[H][m])
        B[P][c] = C.get(P, {}).get(c, B[P][c])
        B[P][b] = C.get(P, {}).get(b, B[P][b])
        B[P][N] = decrypt(C.get(P, {}).get(N, encrypt(B[P][N])))
        B[P][M] = decrypt(C.get(P, {}).get(M, encrypt(B[P][M])))
        B[K][c] = C.get(K, {}).get(c, B[K][c])
        B[K][b] = C.get(K, {}).get(b, B[K][b])
        B[K][N] = decrypt(C.get(K, {}).get(N, encrypt(B[K][N])))
        B[K][M] = decrypt(C.get(K, {}).get(M, encrypt(B[K][M])))
        B[p] = C.get(p, B[p])
        B[w] = C.get(w, B[w])
        B[ft] = C.get(ft, B[ft])
        B[u] = C.get(u, B[u])
        B["loc_path"] = C.get("loc_path", B["loc_path"])
        urls = C.get("loc_urls")
        if urls is I:
            single = C.get("loc_url")
            urls = [single] if single else B.get("loc_urls", LOC_URLS_DEFAULT)
        B["loc_urls"] = urls
        try:
            save_config(B)
        except E:
            pass
    except E as F_:
        try:
            with x(AM, "a", encoding=k) as G_:
                G_.write(
                    f"[{A9.now().strftime(A6)}] [USER: {AO}] [PC: {AF}] ERROR: Failed to load config.json: {F_}\n"
                )
        except:
            pass
    return B


def save_config(config):
    A_ = config
    C_ = {
        H: {
            v: A_[H][v],
            r: A_[H][r],
            N: encrypt(A_[H][N]),
            M: encrypt(A_[H][M]),
            m: A_[H][m],
        },
        P: {c: A_[P][c], b: A_[P][b], N: encrypt(A_[P][N]), M: encrypt(A_[P][M])},
        K: {c: A_[K][c], b: A_[K][b], N: encrypt(A_[K][N]), M: encrypt(A_[K][M])},
        p: A_.get(p, K),
        w: A_.get(w, SQL_UPDATE_TEMPLATE),
        ft: A_.get(ft, J),
        u: A_.get(u, J),
        "loc_path": A_.get("loc_path", LC_DEFAULT),
        "loc_urls": A_.get("loc_urls", LOC_URLS_DEFAULT),
    }
    try:
        with x(AD, T, encoding=k) as D_:
            Ar.dump(C_, D_, indent=4)
    except E as B_:
        O.showerror(AK, CONFIG_SAVE_FAILED_MSG.format(error=B_))
        try:
            with x(AM, "a", encoding=k) as F_:
                F_.write(
                    f"[{A9.now().strftime(A6)}] [USER: {AO}] [PC: {AF}] ERROR: Failed to save config.json: {B_}\n"
                )
        except:
            pass


LANG_CFG = "language.json"


def load_language_pref():
    try:
        with x(A.path.join(LC, LANG_CFG), "r", encoding=k) as f_:
            return Ar.load(f_).get("language", "auto")
    except E:
        return "auto"


def save_language_pref(lang):
    try:
        A.makedirs(LC, exist_ok=J)
        with x(A.path.join(LC, LANG_CFG), T, encoding=k) as f_:
            Ar.dump({"language": lang}, f_, indent=4)
    except E:
        pass


def _expand_github_tree(url):
    try:
        parts = BP.urlparse(url)
        segments = [seg for seg in parts.path.split("/") if seg]
        if Q(segments) >= 4 and segments[2] == "tree":
            owner, repo, _tree, ref, *sub = segments
            api = (
                f"https://api.github.com/repos/{owner}/{repo}/contents/"
                f"{'/'.join(sub)}?ref={ref}"
            )
            with BN.urlopen(api) as resp:
                items = Ar.loads(resp.read().decode(k))
            return [item.get("download_url") for item in items if item.get("type") == "file"]
    except E:
        return []
    return []


def download_localizations(force=Ay):
    ok = J
    try:
        A.makedirs(LC, exist_ok=J)
    except E:
        return Ay
    for url in LOC_URLS:
        targets = []
        if "github.com" in url and "/tree/" in url:
            targets = _expand_github_tree(url)
            if not targets:
                ok = Ay
                continue
        else:
            targets = [url]
        for u_ in targets:
            fname = A.path.basename(BP.urlparse(u_).path)
            if not fname:
                ok = Ay
                continue
            dest = A.path.join(LC, fname)
            if not force and A.path.exists(dest):
                continue
            try:
                with BN.urlopen(u_) as resp, x(dest, "wb") as f_:
                    f_.write(resp.read())
            except E:
                ok = Ay
    return ok


def load_localization(language=I):
    B = language
    if not B or B == "auto":
        try:
            BO.setlocale(BO.LC_ALL, "")
            B = (BO.getlocale()[0] or "en").split("_")[0]
        except E:
            B = "en"
    C_ = {"pl": "pl.json", "ua": "ua.json", "en": "eng.json"}
    F = C_.get(B.lower(), "eng.json")
    paths = [
        A.path.join(LC, F),
        A.path.join(A.path.dirname(A.path.abspath(__file__)), "Localization", F),
    ]
    for path in paths:
        if A.path.exists(path):
            try:
                with x(path, "r", encoding=k) as D_:
                    return Ar.load(D_)
            except E:
                pass
    return {}


D = load_config()
LC = D.get("loc_path", LC_DEFAULT) or LC_DEFAULT
LOC_URLS = D.get("loc_urls", LOC_URLS_DEFAULT)
LANG_PREF = load_language_pref()
LOC_DL_OK = download_localizations()
LANG = load_localization(LANG_PREF)
LANG_EN = load_localization("en")
NO_FILE_LABEL = LANG.get("no_file", NO_FILE_FALLBACK)
LANGUAGE_TAB_LABEL = LANG.get("language_tab", "Język")
LANGUAGE_LABEL = LANG.get("language_label", "Język:")
LOC_PATH_LABEL = LANG.get("loc_path_label", "Folder lokalizacji:")
LOC_URLS_LABEL = LANG.get("loc_urls_label", "Linki lokalizacji:")
UPDATE_LOC_LABEL = LANG.get("loc_update_label", "Aktualizuj")
LOC_UPDATE_SUCCESS_MSG = LANG.get(
    "loc_update_success", "Zaktualizowano pliki lokalizacyjne"
)
PROCESSING_MSG = LANG.get("processing", PROCESSING_MSG)
PROCESSING_UI_MSG = LANG.get(
    "processing_ui", ">>> Processing, please wait..."
)
OPERATION_TITLE = LANG.get("operation_title", OPERATION_TITLE)
NETWORK_ERROR_MSG = LANG.get("network_error", NETWORK_ERROR_MSG)
PATH_NOT_FOUND_MSG = LANG.get("path_not_found", PATH_NOT_FOUND_MSG)
LOGIN_DATA_ERROR_MSG = LANG.get("login_data_error", LOGIN_DATA_ERROR_MSG)
MISSING_FIELDS_MSG = LANG.get("missing_fields", MISSING_FIELDS_MSG)
INCOMPLETE_DATA_MSG = LANG.get("incomplete_data", INCOMPLETE_DATA_MSG)
NO_DATA_MSG = LANG.get("no_data", NO_DATA_MSG)
CANCEL_LABEL = LANG.get("cancel", CANCEL_LABEL)
SETTINGS_LABEL = LANG.get("settings", SETTINGS_LABEL)
EDIT_LISTS_LABEL = LANG.get("edit_lists", EDIT_LISTS_LABEL)
Ac = LANG.get("save_error", Ac)
AJ = LANG.get("not_in_list", AJ)
AK = LANG.get("error", AK)
OPEN_FOLDER_LABEL = LANG.get("open_folder", "Otwórz folder")
CLEAR_LOG_LABEL = LANG.get("clear_log", "Wyczyść log")
VALUE_NOT_EXISTS_QUESTION = LANG.get(
    "value_not_exists_add_question",
    "Wartość '{value}' nie istnieje na liście dodatków. Dodać do listy?",
)
CHANGE_LANGUAGE_LABEL = LANG.get("change_language", "Zmień język")
LANGUAGE_PROMPT = LANG.get("language_prompt", "Kod języka (pl, ua, eng):")
RESTART_TO_APPLY_LABEL = LANG.get(
    "restart_to_apply", "Uruchom ponownie aplikację, aby zastosować zmiany"
)

CONFIG_SAVE_FAILED_MSG = LANG.get(
    "config_save_failed",
    "Nie udało się zapisać pliku konfiguracyjnego:\n{error}",
)
LIST_CREATE_FAILED_MSG = LANG.get(
    "list_create_failed",
    "Nie udało się utworzyć pliku list.xlsx:\n{error}",
)
LIST_SAVE_FAILED_MSG = LANG.get(
    "list_save_failed",
    "Nie udało się zapisać pliku list.xlsx:\n{error}",
)
LIST_DATA_SAVE_FAILED_MSG = LANG.get(
    "list_data_save_failed",
    "Nie udało się zapisać danych do pliku list.xlsx:\n{error}",
)
FOLDER_OPEN_FAILED_MSG = LANG.get(
    "folder_open_failed",
    "Nie udało się otworzyć folderu:\n{error}",
)
OPERATION_ERRORS_MSG = LANG.get(
    "operation_errors",
    "Operacja zakończyła się z błędami. Sprawdź logi oraz folder kopii zapasowej: {backup}",
)
FTP_SEND_FAILED_MSG = LANG.get(
    "ftp_send_failed",
    "Dane lokalne zostały zapisane, jednak wysyłanie plików na serwer FTP nie powiodło się.\nPowód: {reason}",
)
FTP_SKIPPED_NO_EAN_MSG = LANG.get(
    "ftp_skipped_no_ean",
    "Dane lokalne zostały zapisane, jednak nie wysłano zdjęć na FTP z powodu braku prawidłowego kodu EAN-13.",
)
SQL_UPDATE_FAILED_MSG = LANG.get(
    "sql_update_failed",
    "Dane lokalne oraz FTP zostały zaktualizowane, jednak wystąpił błąd podczas aktualizacji bazy danych.\nPowód: {reason}",
)
SAVED_LABEL = LANG.get("saved", "Zapisano")
UPDATE_SUCCESS_MSG = LANG.get(
    "update_success", "Zaktualizowano dane dla EAN {ean}."
)
NO_EAN_LABEL = LANG.get("no_ean", "Brak EAN")
ENTER_EAN_TO_LOAD_MSG = LANG.get(
    "enter_ean_to_load", "Wprowadź kod EAN, aby wczytać dane."
)
CANNOT_SEARCH_NO_EAN_MSG = LANG.get(
    "cannot_search_no_ean", "Nie można wyszukać danych dla 'BRAK-EAN'."
)
NOT_FOUND_LABEL = LANG.get("not_found", "Nie znaleziono")
NO_SAVED_DATA_FOR_EAN_MSG = LANG.get(
    "no_saved_data_for_ean", "Brak zapisanych danych dla EAN {ean}."
)
FILL_REQUIRED_BEFORE_OPEN_MSG = LANG.get(
    "fill_required_before_open",
    "Uzupełnij wszystkie wymagane pola (nazwa, typ, model, kolor 1) przed otwarciem folderu.",
)
CHANGE_DATA_ADMIN_LABEL = LANG.get(
    "change_data_admin", "Zmień dane (Administrator)"
)
DATABASE_LABEL = LANG.get("database_label", "Baza danych:")
SERVER_LABEL = LANG.get("server_label", "Serwer:")
MSSQL_SERVER_LABEL = LANG.get("mssql_server", "MS SQL Server")
TEST_BUTTON_LABEL = LANG.get("test_button", "Testuj")
CONNECTED_LABEL = LANG.get("connected", "Połączono")
PASSWORD_LABEL = LANG.get("password_label", "Hasło:")
USER_LABEL = LANG.get("user_label", "Użytkownik:")
MYSQL_LABEL = LANG.get("mysql_label", "MySQL")
SAVE_LABEL = LANG.get("save", "Zapisz")
NO_PERMISSIONS_LABEL = LANG.get("no_permissions", "Brak uprawnień")
RUN_AS_ADMIN_MSG = LANG.get(
    "run_as_admin",
    "Uruchom operację z uprawnieniami administratora, aby edytować te ustawienia.",
)
IMAGE_SETTINGS_LABEL = LANG.get(
    "image_settings", "Ustawienia przetwarzania obrazów:"
)
RESIZE_LABEL = LANG.get(
    "resize_label", "Zmniejszaj obrazy większe niż"
)
PX_MAX_LABEL = LANG.get("px_max_label", "px (max wymiar)")
COMPRESS_LABEL = LANG.get(
    "compress_label", "Kompresuj JPEG (jakość)"
)
LIMIT_SIZE_LABEL = LANG.get(
    "limit_size_label", "Ogranicz rozmiar pliku do"
)
CONVERT_TIF_LABEL = LANG.get(
    "convert_tif_label", "Konwertuj .tif na"
)
FTP_SERVER_LABEL = LANG.get("ftp_server_label", "Serwer FTP:")
PORT_LABEL = LANG.get("port_label", "Port:")
FTP_PATH_LABEL = LANG.get(
    "ftp_path_label", "Ścieżka (katalog) na serwerze:"
)
FTP_TEST_LABEL = LANG.get(
    "ftp_test_label", "Test połączenia FTP:"
)
FTP_UPDATE_LABEL = LANG.get(
    "ftp_update_label", "Aktualizuj pliki na FTP:"
)
DB_TYPE_LABEL = LANG.get("db_type_label", "Typ bazy danych:")
SQL_UPDATE_LABEL = LANG.get(
    "sql_update_label", "Aktualizuj bazę przy zapisie:"
)
SQL_QUERY_LABEL = LANG.get("sql_query_label", "Zapytanie SQL:")
SQL_TEST_LABEL = LANG.get("sql_test_label", "Test połączenia SQL:")
NAME_LABEL = LANG.get("name_label", "Nazwa mebla*:")
TYPE_LABEL = LANG.get("type_label", "Typ mebla*:")
MODEL_LABEL = LANG.get("model_label", "Model mebla*:")
COLOR1_LABEL = LANG.get("color1_label", "Kolor 1*:")
COLOR2_LABEL = LANG.get("color2_label", "Kolor 2:")
COLOR3_LABEL = LANG.get("color3_label", "Kolor 3:")
EXTRA_LABEL = LANG.get("extra_label", "Dodatkowe:")
EAN_OPTIONAL_LABEL = LANG.get(
    "ean_optional_label", "EAN (opcjonalnie):"
)
LOAD_LABEL = LANG.get("load_label", "Wczytaj")
UPDATE_LABEL = LANG.get("update_label", "Aktualizuj")
CHOOSE_LABEL = LANG.get("choose_label", "Wybierz")
NEW_COMBINATION_LABEL = LANG.get("new_combination_label", "Nowa kombinacja")
FTP_ERROR_LABEL = LANG.get("ftp_error", "Błąd FTP")
SQL_ERROR_LABEL = LANG.get("sql_error", "Błąd SQL")
IMAGES_TAB_LABEL = LANG.get("images_tab", "Obrazy")
FTP_TAB_LABEL = LANG.get("ftp_tab", "FTP")
SQL_TAB_LABEL = LANG.get("sql_tab", "SQL")
WARNING_LABEL = LANG.get("warning", "Uwaga")
SELECT_COMBINATION_TITLE = LANG.get(
    "select_combination_title", "Wybierz istniejącą kombinację"
)
SELECT_COMBINATION_PROMPT = LANG.get(
    "select_combination_prompt",
    "Wybierz istniejącą kombinację kolorów:",
)
SELECT_FILE_TITLE = LANG.get("select_file_title", "Wybierz plik")
OTHER_ERROR_MSG = LANG.get("other_error", "Inny błąd: {error}")
FTP_GENERIC_ERROR_MSG = LANG.get("ftp_generic_error", "Błąd FTP: {error}")
FILL_REQUIRED_BEFORE_SUBMIT_MSG = LANG.get(
    "fill_required_before_submit",
    "Uzupełnij wszystkie wymagane pola oznaczone * przed zatwierdzeniem.",
)
EAN_PROMPT_TITLE = LANG.get("ean_prompt_title", "EAN")
EAN_MISSING_PROMPT = LANG.get(
    "ean_missing_prompt",
    "Nie podano EAN.\nWprowadź kod EAN (13 cyfr) lub pozostaw puste aby użyć 'BRAK-EAN':",
)
APP_TITLE = LANG.get("app_title", "Katalogowanie zdjęć mebli")


def rotate_log(path, max_bytes=1073741824, backups=3):
    B_ = path
    try:
        if A.path.exists(B_) and A.path.getsize(B_) >= max_bytes:
            for C in Ax(backups, 0, -1):
                F = f"{B_}.{C}" if C > 1 else B_
                D_ = f"{B_}.{C+1}"
                if A.path.exists(F):
                    try:
                        if A.path.exists(D_):
                            A.remove(D_)
                    except:
                        pass
                    try:
                        A.rename(F, D_)
                    except:
                        pass
            with x(B_, T, encoding=k) as G_:
                G_.write(f"[{A9.now().strftime(A6)}] Log rotated\n")
    except E:
        pass


def log_error(message, ui_message=None):
    A_ = message
    try:
        rotate_log(AM)
        B_ = A9.now().strftime(A6)
        with x(AM, "a", encoding=k) as C_:
            C_.write(f"[{B_}] [USER: {AO}] [PC: {AF}] ERROR: {A_}\n")
    except E:
        pass
    try:
        if AG:
            if threading.current_thread() != threading.main_thread():
                AG.after(0, lambda msg=(ui_message or A_): AG._ui_log(f"❗ {msg}"))
            else:
                AG._ui_log(f"❗ {ui_message or A_}")
    except E:
        pass


def log_info(message, ui_message=None):
    A_ = message
    try:
        rotate_log(BM)
        B_ = A9.now().strftime(A6)
        with x(BM, "a", encoding=k) as C_:
            C_.write(f"[{B_}] [USER: {AO}] [PC: {AF}] {A_}\n")
    except E:
        pass
    try:
        if AG:
            if threading.current_thread() != threading.main_thread():
                AG.after(0, lambda msg=(ui_message or A_): AG._ui_log(f"• {msg}"))
            else:
                AG._ui_log(f"• {ui_message or A_}")
    except E:
        pass


def log_error_loc(key, **kwargs):
    file_msg = LANG_EN.get(key, key).format(**kwargs)
    ui_msg = LANG.get(key, file_msg).format(**kwargs)
    log_error(file_msg, ui_msg)


def log_info_loc(key, **kwargs):
    file_msg = LANG_EN.get(key, key).format(**kwargs)
    ui_msg = LANG.get(key, file_msg).format(**kwargs)
    log_info(file_msg, ui_msg)


def is_admin():
    try:
        if A.name == "nt":
            B_ = ctypes.windll.shell32.ShellExecuteW(
                AQ, "runas", "cmd.exe", "/c exit", AQ, 1
            )
            return B_ > 32
        else:
            return Al
    except E:
        return Ay


def get_file_lock_user(path):
    I_ = "latin-1"
    D_ = "ignore"
    F_ = path
    if not A.path.exists(F_):
        return h
    try:
        P = A.open(F_, A.O_RDWR | A.O_EXCL)
        A.close(P)
        return h
    except Au:
        R_ = A.path.dirname(F_)
        K_ = A.path.basename(F_)
        L_ = A.path.join(R_, "~$" + K_)
        if A.path.exists(L_):
            try:
                with x(L_, "rb") as S:
                    C = S.read()
                    if Q(C) >= 2:
                        M = C[1]
                        if 2 + M <= Q(C):
                            N = C[2 : 2 + M]
                            try:
                                H = N.decode(k, errors=D_).strip()
                            except:
                                H = N.decode(I_, errors=D_).strip()
                            if H:
                                return H
                    G_ = C.decode(k, errors=D_)
                    if not G_ or G_.count("\x00") > 0:
                        G_ = C.decode(I_, errors=D_)
                    T = G_.replace(K_, B)
                    O_ = [A for A in T.split() if 3 <= Q(A) <= 50]
                    if O_:
                        return max(O_, key=Q)
            except E:
                pass
        return J


Aw = [
    ("01", "Assembly_instruction"),
    ("02", "Assembly_instruction1"),
    ("03", "DETAIL_pic"),
    ("04", "DETAIL_pic1"),
    ("05", "element_pic1"),
    ("06", ELEMENT_PIC),
    ("07", "LED_Assembly_instruction"),
    ("08", "MOOD_pic"),
    ("09", "MOOD_pic1"),
    ("10", "MOOD_pic2"),
    ("11", "MOOD_pic3"),
    ("12", "MOOD_pic4"),
    ("13", "MOOD_pic5"),
    ("14", NON_PIC),
    ("15", OPEN_FURNITURE),
    ("16", "open_furniture1"),
    ("17", "open_furniture2"),
    ("18", "NO_EAN"),
    ("19", "Technical_drawing"),
    ("20", "Technical_drawing1"),
    ("21", "Technical_drawing2"),
    ("22", "WB_pic"),
    ("23", "WB_pic1"),
    ("24", "WB_pic2"),
    ("25", "WB_pic3"),
    ("26", "WB_pic4"),
]


def label_category(label):
    B_ = label.rstrip("0123456789")
    if B_.startswith("LED_"):
        B_ = B_[4:]
    A_ = B_.lower()
    if "assembly_instruction" in A_:
        return "ASSEMBLY"
    elif "technical_drawing" in A_:
        return "TECHNICAL"
    elif "mood_pic" in A_:
        return "MOOD"
    elif "wb_pic" in A_:
        return "WB"
    elif "detail_pic" in A_:
        return "DETAIL"
    elif ELEMENT_PIC in A_:
        return "ELEMENT"
    elif OPEN_FURNITURE in A_:
        return "OPEN-FURNITURE"
    elif "no_ean" in A_:
        return "NO-EAN"
    elif NON_PIC in A_:
        return "NON-PIC"
    else:
        return B_.replace(a, g).upper()


def prepare_excel_lists():
    if not A.path.isdir(A.path.dirname(o)):
        A.makedirs(A.path.dirname(o), exist_ok=J)
    if not A.path.exists(o):
        F = BV()
        F.remove(F.active)
        for D_ in AE.values():
            K = F.create_sheet(title=D_)
            if D_ == W:
                K.append(["EAN", Ae, Ad, AZ, AY, AX, AW, d])
        try:
            F.save(o)
        except E as V_:
            O.showerror(AK, LIST_CREATE_FAILED_MSG.format(error=V_))
            log_error_loc("excel_create_failed", error=V_)
    F = Ah(o)
    L_ = {}
    for D_ in AE.values():
        K = F[D_]
        if D_ == W:
            X_ = {}
            for C_ in K.iter_rows(min_row=2, values_only=J):
                if not C_[0]:
                    continue
                Z_ = G(C_[0]).strip()
                M_ = G(C_[1]) if C_[1] else B
                N_ = G(C_[2]) if C_[2] else B
                P_ = G(C_[3]) if C_[3] else B
                Q_ = G(C_[4]) if C_[4] else B
                R_ = G(C_[5]) if C_[5] else B
                S_ = G(C_[6]) if C_[6] else B
                I_ = G(C_[7]) if C_[7] else B
                M_ = M_.strip().upper()
                N_ = N_.strip().upper()
                P_ = P_.strip().upper()
                Q_ = Q_.strip().upper()
                R_ = R_.strip().upper()
                S_ = S_.strip().upper()
                I_ = I_.strip()
                I_ = I_.replace(a, g).upper()
                X_[Z_] = {Ae: M_, Ad: N_, AZ: P_, AY: Q_, AX: R_, AW: S_, d: I_}
            L_[D_] = X_
        else:
            T = []
            for Y_ in K["A"]:
                if Y_.value:
                    H = G(Y_.value).strip()
                    if D_ == d:
                        H = H.replace(a, g)
                    H = H.upper()
                    if H not in T:
                        T.append(H)
            L_[D_] = T
    return L_


def add_to_list(sheet_name, value):
    B_ = sheet_name
    A_ = value
    if not A_:
        return
    A_ = G(A_).strip().upper()
    if B_ == AE[d]:
        A_ = A_.replace(a, g)
    C_ = get_file_lock_user(o)
    if C_:
        D_ = f"przez użytkownika '{C_}'" if Aq(C_, G) else Ap
        O.showerror(
            Ao,
            LANG.get(
                "excel_file_open",
                f"Nie można zapisać listy. Plik Excel jest otwarty {D_}. Zamknij plik i spróbuj ponownie.",
            ).format(reason=D_),
        )
        log_error_loc(
            "excel_add_locked", value=A_, list=B_, reason=D_
        )
        return
    F = Ah(o)
    H = F[B_]
    J_ = [G(A.value).strip().upper() for A in H["A"] if A.value]
    if A_ not in J_:
        H.append([A_])
        try:
            F.save(o)
        except E as I_:
            O.showerror(Ac, LIST_SAVE_FAILED_MSG.format(error=I_))
            log_error_loc(
                "excel_add_save_failed", value=A_, list=B_, error=I_
            )
            return
        log_info_loc("list_value_added", value=A_, list=B_)


def remove_from_list(sheet_name, value):
    A_ = value
    B_ = sheet_name
    C_ = get_file_lock_user(o)
    if C_:
        F_ = f"przez użytkownika '{C_}'" if Aq(C_, G) else Ap
        O.showerror(
            Ao,
            LANG.get(
                "excel_file_open",
                f"Nie można zapisać listy. Plik Excel jest otwarty {F_}. Zamknij plik i spróbuj ponownie.",
            ).format(reason=F_),
        )
        log_error_loc(
            "excel_remove_locked", value=A_, list=B_, reason=F_
        )
        return
    H = Ah(o)
    I_ = H[B_]
    K_ = G(A_).strip().upper()
    for D_ in I_["A"]:
        if D_.value and G(D_.value).strip().upper() == K_:
            I_.delete_rows(D_.row, 1)
            break
    try:
        H.save(o)
    except E as J_:
        O.showerror(Ac, LIST_SAVE_FAILED_MSG.format(error=J_))
        log_error_loc(
            "excel_remove_save_failed", value=A_, list=B_, error=J_
        )
        return
    log_info_loc("list_value_removed", value=A_, list=B_)


def save_ean_entry(ean, name, typ, model, col1, col2, col3, extra):
    R = ean
    K_ = col1
    M_ = model
    N_ = typ
    P_ = name
    F_ = col3
    H_ = col2
    X_ = get_file_lock_user(o)
    if X_:
        i_ = f"przez użytkownika '{X_}'" if Aq(X_, G) else Ap
        O.showerror(
            Ao,
            LANG.get(
                "excel_data_file_open",
                f"Nie można zapisać danych. Plik Excel jest otwarty {i_}. Zamknij plik i spróbuj ponownie.",
            ).format(reason=i_),
        )
        log_error_loc(
            "excel_entry_save_locked", ean=R, reason=i_
        )
        return h
    j = Ah(o)
    V = j[AE[W]]
    P_ = G(P_).strip().upper()
    N_ = G(N_).strip().upper()
    M_ = G(M_).strip().upper()
    K_ = G(K_).strip().upper()
    H_ = G(H_).strip().upper() if H_ else B
    F_ = G(F_).strip().upper() if F_ else B
    D_ = G(extra).strip()
    if D_ == B or D_.upper() in [L, L]:
        D_ = L
    else:
        D_ = D_.replace(a, g).upper()
    Y = h
    Q_ = I
    for A_ in V.iter_rows(min_row=2):
        T = A_[0].value
        if T is I:
            continue
        if G(T).upper() == G(R).upper():
            Q_ = A_
            Y = J
            break
    if Q_:
        Q_[1].value = P_
        Q_[2].value = N_
        Q_[3].value = M_
        Q_[4].value = K_
        Q_[5].value = H_ or B
        Q_[6].value = F_ or B
        Q_[7].value = D_
    else:
        l_ = G(R).strip()
        if l_.upper() != q:
            C_ = I
            for A_ in V.iter_rows(min_row=2):
                T = G(A_[0].value).strip().upper() if A_[0].value else B
                if T == q:
                    Z_ = G(A_[1].value).strip().upper() if A_[1].value else B
                    b_ = G(A_[2].value).strip().upper() if A_[2].value else B
                    c_ = G(A_[3].value).strip().upper() if A_[3].value else B
                    d_ = G(A_[4].value).strip().upper() if A_[4].value else B
                    e_ = G(A_[5].value).strip().upper() if A_[5].value else B
                    f_ = G(A_[6].value).strip().upper() if A_[6].value else B
                    S_ = G(A_[7].value).strip() if A_[7].value else B
                    S_ = S_.replace(a, g).upper()
                    if (
                        Z_ == P_
                        and b_ == N_
                        and c_ == M_
                        and d_ == K_
                        and e_ == (H_ or B)
                        and f_ == (F_ or B)
                        and S_ == D_
                    ):
                        C_ = A_
                        Y = J
                        break
            if C_:
                C_[0].value = G(R)
                C_[1].value = P_
                C_[2].value = N_
                C_[3].value = M_
                C_[4].value = K_
                C_[5].value = H_ or B
                C_[6].value = F_ or B
                C_[7].value = D_
            else:
                V.append([G(R), P_, N_, M_, K_, H_ or B, F_ or B, D_])
        else:
            C_ = I
            for A_ in V.iter_rows(min_row=2):
                T = G(A_[0].value).strip().upper() if A_[0].value else B
                if T == q:
                    Z_ = G(A_[1].value).strip().upper() if A_[1].value else B
                    b_ = G(A_[2].value).strip().upper() if A_[2].value else B
                    c_ = G(A_[3].value).strip().upper() if A_[3].value else B
                    d_ = G(A_[4].value).strip().upper() if A_[4].value else B
                    e_ = G(A_[5].value).strip().upper() if A_[5].value else B
                    f_ = G(A_[6].value).strip().upper() if A_[6].value else B
                    S_ = G(A_[7].value).strip() if A_[7].value else B
                    S_ = S_.replace(a, g).upper()
                    if (
                        Z_ == P_
                        and b_ == N_
                        and c_ == M_
                        and d_ == K_
                        and e_ == (H_ or B)
                        and f_ == (F_ or B)
                        and S_ == D_
                    ):
                        C_ = A_
                        Y = J
                        break
            if C_:
                C_[1].value = P_
                C_[2].value = N_
                C_[3].value = M_
                C_[4].value = K_
                C_[5].value = H_ or B
                C_[6].value = F_ or B
                C_[7].value = D_
            else:
                V.append([G(R), P_, N_, M_, K_, H_ or B, F_ or B, D_])
    try:
        j.save(o)
    except E as k_:
        O.showerror(Ac, LIST_DATA_SAVE_FAILED_MSG.format(error=k_))
        log_error_loc(
            "excel_entry_save_failed", ean=R, error=k_
        )
        return h
    return J


def connect_db():
    C_ = D.get(p, K).lower()
    if C_ == K:
        A_ = D[K]
        return mysql.connector.connect(
            host=A_[c],
            user=A_[N],
            password=A_[M],
            database=A_[b],
            connection_timeout=5,
            use_pure=True,
        )
    A_ = D[P]
    F_ = A_.get(c)
    G_ = A_.get(b)
    H_ = A_.get(N)
    J_ = A_.get(M)
    last_exc = None
    extra = "Encrypt=yes;TrustServerCertificate=yes;Connection Timeout=5"
    try:
        drivers_seen = pyodbc.drivers()
    except E:
        drivers_seen = []
    for L_ in BW:
        try:
            O_ = f"DRIVER={{{L_}}};SERVER={F_};DATABASE={G_};UID={H_};PWD={J_};{extra}"
            return pyodbc.connect(O_)
        except E as ex:
            last_exc = ex
            continue
    import struct

    arch = f"{struct.calcsize("P")*8}-bit EXE on {BR.platform()}"
    msg = f"Brak działającego sterownika ODBC do MSSQL.\nPróbowano: {", ".join(BW)}\nSystem widzi sterowniki: {", ".join(drivers_seen)or"(brak)"}\nArchitektura: {arch}\nOstatni błąd: {last_exc}"
    raise E(msg)


class App(BU.Tk):
    def __init__(B):
        super().__init__()
        B.title(APP_TITLE)
        B.geometry("1200x800")
        B.style = C.Style()
        B.style.theme_use("clam")
        B.style.configure(Z, fieldbackground=LIGHT_GREEN)
        D_ = prepare_excel_lists()
        B.entries = D_.get(W, {})
        if W in D_:
            D_.pop(W)
        B.lists = D_
        if not A.path.isdir(l):
            A.makedirs(l, exist_ok=J)
        E_ = [B_.upper() for B_ in A.listdir(l) if A.path.isdir(A.path.join(l, B_))]
        G_ = [A_ for A_ in B.lists[n] if A_ not in E_]
        B.lists[n] = E_ + G_
        B.var_name = F.StringVar()
        B.var_type = F.StringVar()
        B.var_model = F.StringVar()
        B.var_color1 = F.StringVar()
        B.var_color2 = F.StringVar()
        B.var_color3 = F.StringVar()
        B.var_extra = F.StringVar()
        B.var_ean = F.StringVar()
        B.pending_additions = {}
        B.pending_deletions = {}
        B.pending_ftp_deletions = {}
        B.ftp_remote_only = {}
        B.ftp_presence = {}
        B.ftp_downloaded_final = set()
        B.opt_resize = F.BooleanVar(value=J)
        B.opt_compress = F.BooleanVar(value=h)
        B.opt_maxsize = F.BooleanVar(value=h)
        B.resize_max_dim = F.IntVar(value=2000)
        B.compress_quality = F.IntVar(value=85)
        B.max_file_kb = F.IntVar(value=500)
        B.opt_convert_tif = F.BooleanVar(value=J)
        B.tif_target_format = F.StringVar(value=At)
        B.loading_by_ean = h
        B.suppress_scan = h
        B.model_select_win_open = h
        B.dragging_idx = I
        B.original_files = {}
        B.is_processing = h
        B.logged_counts = h
        B.suppress_next_lookup = h
        B._build_form()
        B._build_slots()
        H_ = Q(E_)
        B.combo_name.existing_count = H_
        set_app(B)

    def _build_form(A):
        F_ = "<FocusOut>"
        D_ = "<KeyRelease>"
        E_ = "<Return>"
        B_ = C.Frame(A)
        B_.pack(side="top", fill="x", padx=10, pady=10)
        G_ = C.Label(B_, text=NAME_LABEL)
        G_.grid(row=0, column=0, sticky=R)
        A._add_tooltip(
            G_,
            LANG.get(
                "name_tooltip",
                "Pełna nazwa mebla bez kolorów, typu i modelu, np: 'Maggiore', 'LUNA', 'SLANT'.",
            ),
        )
        A.combo_name = C.Combobox(
            B_, textvariable=A.var_name, values=A.lists[n], state=X
        )
        A.combo_name.grid(row=0, column=1, padx=5, pady=2)
        A.combo_name.bind(E_, lambda e: A._on_name_commit())
        A.combo_name.bind(A2, lambda e: A._on_name_commit())
        A.combo_name.bind(F_, lambda e: A._on_name_commit())
        A.combo_name.bind(D_, A._on_key_release)
        H_ = C.Label(B_, text=TYPE_LABEL)
        H_.grid(row=1, column=0, sticky=R)
        A._add_tooltip(
            H_,
            LANG.get(
                "type_tooltip",
                "Typ mebla, np: 'KOMODA', 'RTV', 'STÓŁ' (można dodać długość, np. 'RTV 100', 'SZAFA 80').",
            ),
        )
        A.combo_type = C.Combobox(
            B_, textvariable=A.var_type, values=A.lists[t], state=V
        )
        A.combo_type.grid(row=1, column=1, padx=5, pady=2)
        A.combo_type.bind(E_, lambda e: A._on_type_commit())
        A.combo_type.bind(A2, lambda e: A._on_type_commit())
        A.combo_type.bind(F_, lambda e: A._on_type_commit())
        A.combo_type.bind(D_, A._on_key_release)
        I_ = C.Label(B_, text=MODEL_LABEL)
        I_.grid(row=2, column=0, sticky=R)
        A._add_tooltip(
            I_,
            LANG.get(
                "model_tooltip",
                "Model lub wersja mebla, np: 'MA03', 'Li01', 'SOL-05'.",
            ),
        )
        A.combo_model = C.Combobox(
            B_, textvariable=A.var_model, values=A.lists[s], state=V
        )
        A.combo_model.grid(row=2, column=1, padx=5, pady=2)
        A.combo_model.bind(E_, lambda e: A._on_model_commit())
        A.combo_model.bind(A2, lambda e: A._on_model_commit())
        A.combo_model.bind(D_, A._on_key_release)
        J_ = C.Label(B_, text=COLOR1_LABEL)
        J_.grid(row=3, column=0, sticky=R)
        A._add_tooltip(
            J_, LANG.get("color1_tooltip", "Główny kolor mebla (wymagany).")
        )
        A.combo_color1 = C.Combobox(
            B_, textvariable=A.var_color1, values=A.lists[Y], state=V
        )
        A.combo_color1.grid(row=3, column=1, padx=5, pady=2)
        A.combo_color1.bind(E_, lambda e: A._on_color_commit())
        A.combo_color1.bind(A2, lambda e: A._on_color_commit())
        A.combo_color1.bind(F_, lambda e: A._on_color_commit())
        A.combo_color1.bind(D_, A._on_key_release)
        K_ = C.Label(B_, text=COLOR2_LABEL)
        K_.grid(row=4, column=0, sticky=R)
        A._add_tooltip(
            K_, LANG.get("color2_tooltip", "Drugi kolor mebla (opcjonalnie).")
        )
        A.combo_color2 = C.Combobox(
            B_, textvariable=A.var_color2, values=A.lists[Y], state=V
        )
        A.combo_color2.grid(row=4, column=1, padx=5, pady=2)
        A.combo_color2.bind(E_, lambda e: A._on_color_commit())
        A.combo_color2.bind(A2, lambda e: A._on_color_commit())
        A.combo_color2.bind(F_, lambda e: A._on_color_commit())
        A.combo_color2.bind(D_, A._on_key_release)
        L_ = C.Label(B_, text=COLOR3_LABEL)
        L_.grid(row=5, column=0, sticky=R)
        A._add_tooltip(
            L_, LANG.get("color3_tooltip", "Trzeci kolor mebla (opcjonalnie).")
        )
        A.combo_color3 = C.Combobox(
            B_, textvariable=A.var_color3, values=A.lists[Y], state=V
        )
        A.combo_color3.grid(row=5, column=1, padx=5, pady=2)
        A.combo_color3.bind(E_, lambda e: A._on_color_commit())
        A.combo_color3.bind(A2, lambda e: A._on_color_commit())
        A.combo_color3.bind(F_, lambda e: A._on_color_commit())
        A.combo_color3.bind(D_, A._on_key_release)
        M_ = C.Label(B_, text=EXTRA_LABEL)
        M_.grid(row=6, column=0, sticky=R)
        A._add_tooltip(
            M_,
            LANG.get(
                "extra_tooltip",
                "Dodatkowe informacje, np. LED, RGB (pozostaw puste, jeśli brak dodatków).",
            ),
        )
        A.combo_extra = C.Combobox(
            B_, textvariable=A.var_extra, values=A.lists[d], state=V
        )
        A.combo_extra.grid(row=6, column=1, padx=5, pady=2)
        A.combo_extra.bind(E_, lambda e: A._on_extra_commit())
        A.combo_extra.bind(A2, lambda e: A._on_extra_commit())
        A.combo_extra.bind(F_, lambda e: A._on_extra_commit())
        A.combo_extra.bind(D_, A._on_key_release)
        N_ = C.Label(B_, text=EAN_OPTIONAL_LABEL)
        N_.grid(row=7, column=0, sticky=R)
        A._add_tooltip(
            N_,
            LANG.get(
                "ean_tooltip",
                "13-cyfrowy kod EAN produktu. Jeśli nie podany, zostanie użyte 'BRAK-EAN'.",
            ),
        )
        A.entry_ean = C.Entry(B_, textvariable=A.var_ean, state=X)
        A.entry_ean.grid(row=7, column=1, padx=5, pady=2)
        O_ = C.Button(B_, text=LOAD_LABEL, command=A._load_by_ean)
        O_.grid(row=7, column=2, padx=5, pady=2)
        P_ = C.Button(B_, text=EDIT_LISTS_LABEL, command=A._open_list_editor)
        P_.grid(row=0, column=2, padx=20)
        Q_ = C.Button(B_, text=SETTINGS_LABEL, command=A._open_settings)
        Q_.grid(row=0, column=3, padx=5)
        A.btn_submit = C.Button(B_, text=UPDATE_LABEL, command=A._on_submit)
        A.btn_submit.grid(row=8, column=0, columnspan=2, pady=10)
        A.btn_open = C.Button(B_, text=OPEN_FOLDER_LABEL, command=A._open_current_folder)
        A.btn_open.grid(row=8, column=2, padx=5, pady=10)
        A.ui_log = BS.ScrolledText(B_, width=48, height=8, state=Ak, wrap="word")
        A.ui_log.grid(row=0, column=4, rowspan=9, padx=10, sticky="nsew")
        S_ = C.Button(B_, text=CLEAR_LOG_LABEL, command=lambda: A._ui_log(clear=Al))
        S_.grid(row=8, column=3, padx=5, pady=10, sticky="e")
        B_.grid_columnconfigure(4, weight=1)

    def _build_slots(B):
        Q_ = "<Button-1>"
        R_ = "#ddd"
        S_ = "<Configure>"
        L_ = "units"
        M_ = C.Frame(B)
        M_.pack(fill=z, expand=J, padx=10, pady=10)
        A_ = F.Canvas(M_)
        T = C.Scrollbar(M_, orient=An, command=A_.yview)
        N_ = C.Frame(A_)
        N_.bind(S_, lambda e: A_.configure(scrollregion=A_.bbox("all")))
        Y = A_.create_window((0, 0), window=N_, anchor="nw")
        A_.bind(S_, lambda e, cw=Y: A_.itemconfig(cw, width=e.width))
        A_.configure(yscrollcommand=T.set)
        A_.pack(side=Am, fill=z, expand=J)
        T.pack(side=AV, fill="y")
        A_.bind_all(
            "<MouseWheel>", lambda e: A_.yview_scroll(int(-1 * (e.delta / 120)), L_)
        )
        A_.bind_all("<Button-4>", lambda e: A_.yview_scroll(-1, L_))
        A_.bind_all("<Button-5>", lambda e: A_.yview_scroll(1, L_))
        B.slots_frame = N_
        B.slots = []
        U = 5
        for G_, (V_, W_) in A0(Aw):
            Z_, O_ = divmod(G_, U)
            H_ = F.Frame(
                B.slots_frame,
                highlightthickness=0,
                highlightbackground=A8,
                highlightcolor=A8,
                bd=0,
            )
            H_.grid(row=Z_, column=O_, padx=5, pady=5, sticky="nsew")
            C.Label(H_, text=f"{V_} {W_}").pack()
            E_ = F.Frame(H_, height=100, bg=R_)
            E_.pack_propagate(h)
            E_.pack(fill=z, expand=J, padx=5, pady=5)
            D_ = F.Label(E_, text=NO_FILE_LABEL, bg=R_)
            D_.pack(fill=z, expand=J)
            D_.drop_target_register(DND_ALL)
            D_.dnd_bind("<<Drop>>", lambda e, i=G_: B._on_drop(e, i))
            K_ = F.Label(E_, text="✕", fg=AT, bg=Ab)
            K_.bind(Q_, lambda e, i=G_: B._remove_file(i))
            K_.place(relx=0, rely=0, anchor="nw")
            K_.place_forget()
            X_ = F.Label(E_, text="...", fg=AT, bg="black")
            X_.bind(Q_, lambda e, i=G_: B._select_file(i))
            X_.place(relx=1.0, rely=0, anchor="ne")
            local_icon = F.Canvas(
                E_,
                width=30,
                height=20,
                highlightthickness=0,
                bd=1,
                relief="solid",
            )
            local_icon.create_text(15, 10, text="LOCAL", font=("Arial", 7), fill="white")
            local_icon.offset_x = -30
            local_icon.place(relx=1.0, rely=1.0, anchor="se", x=local_icon.offset_x)
            local_icon.place_forget()
            ftp_icon = F.Canvas(
                E_,
                width=30,
                height=20,
                highlightthickness=0,
                bd=1,
                relief="solid",
            )
            ftp_icon.create_text(15, 10, text="FTP", font=("Arial", 7), fill="white")
            ftp_icon.offset_x = 0
            ftp_icon.place(relx=1.0, rely=1.0, anchor="se", x=ftp_icon.offset_x)
            ftp_icon.place_forget()
            D_.drag_source_register(1, BJ)
            D_.dnd_bind("<<DragInitCmd>>", lambda e, i=G_: B._on_drag_init(e, i))
            D_.dnd_bind("<<DragEndCmd>>", lambda e: B._on_drag_end(e))
            B.slots.append({Aa: V_, "label": W_, y: D_, A7: K_, "local_icon": local_icon, "ftp_icon": ftp_icon, f: I, AS: H_, B0: I})
        for O_ in Ax(U):
            B.slots_frame.columnconfigure(O_, weight=1)

    def _set_icon_status(C, icon, present):
        if not icon:
            return
        icon.place(relx=1.0, rely=1.0, anchor="se", x=getattr(icon, "offset_x", 0))
        icon.config(bg="green" if present else "red")

    def _refresh_combobox_list(B, combobox, all_values, existing_count=0):
        A_ = combobox
        A_[S] = all_values
        A_.existing_count = existing_count

    def _on_name_commit(C):
        D_ = C.var_name.get().strip()
        if not D_:
            return
        if D_.upper() not in C.lists[n]:
            if O.askyesno(
                AJ, f"Nazwa '{D_}' nie istnieje na liście. Czy dodać ją do listy?"
            ):
                H = C._open_list_editor(n)
                C.wait_window(H)
                C.lists = prepare_excel_lists()
                C.entries = C.lists.get(W, {})
                if W in C.lists:
                    C.lists.pop(W)
                C.combo_name[S] = C.lists[n]
                if D_.upper() not in [A.upper() for A in C.lists[n]]:
                    C.var_name.set(B)
                    return
            else:
                C.var_name.set(B)
                return
        F = A.path.join(l, D_.upper())
        E_ = []
        if A.path.isdir(F):
            E_ = [B for B in A.listdir(F) if A.path.isdir(A.path.join(F, B))]
            C.combo_name.configure(style=Z)
        else:
            C.combo_name.configure(style=j)
        I = [A for A in C.lists[t] if A not in E_]
        C._refresh_combobox_list(C.combo_type, E_ + I, existing_count=Q(E_))
        C.combo_type.configure(state=X)
        C.var_type.set(B)
        C.var_model.set(B)
        C.var_color1.set(B)
        C.var_color2.set(B)
        C.var_color3.set(B)
        C.var_extra.set(B)
        C.var_ean.set(B)
        for G_ in (
            C.combo_type,
            C.combo_model,
            C.combo_color1,
            C.combo_color2,
            C.combo_color3,
            C.combo_extra,
        ):
            G_.configure(style=j)
        for G_ in (
            C.combo_model,
            C.combo_color1,
            C.combo_color2,
            C.combo_color3,
            C.combo_extra,
        ):
            G_.configure(state=V)
        C.btn_submit.configure(state=V)
        C.btn_open.configure(state=V)
        C.entry_ean.configure(state=X)
        C._clear_all_slots()

    def _on_type_commit(C):
        G_ = C.var_name.get().strip()
        D_ = C.var_type.get().strip()
        if not G_ or not D_:
            return
        if D_.upper() not in C.lists[t]:
            if O.askyesno(
                AJ, f"Typ '{D_}' nie istnieje na liście. Czy dodać go do listy?"
            ):
                H = C._open_list_editor(t)
                C.wait_window(H)
                C.lists = prepare_excel_lists()
                C.entries = C.lists.get(W, {})
                if W in C.lists:
                    C.lists.pop(W)
                C.combo_type[S] = C.lists[t]
                if D_.upper() not in [A.upper() for A in C.lists[t]]:
                    C.var_type.set(B)
                    return
            else:
                C.var_type.set(B)
                return
        F = A.path.join(l, G_.upper(), D_.upper())
        E_ = []
        if A.path.isdir(F):
            E_ = [B for B in A.listdir(F) if A.path.isdir(A.path.join(F, B))]
            C.combo_type.configure(style=Z)
        else:
            C.combo_type.configure(style=j)
        I = [A for A in C.lists[s] if A not in E_]
        C._refresh_combobox_list(C.combo_model, E_ + I, existing_count=Q(E_))
        C.combo_model.configure(state=X)
        C.var_model.set(B)
        C.var_color1.set(B)
        C.var_color2.set(B)
        C.var_color3.set(B)
        C.var_extra.set(B)
        C.var_ean.set(B)
        for J_ in (C.combo_color1, C.combo_color2, C.combo_color3, C.combo_extra):
            J_.configure(style=j, state=V)
        C.btn_submit.configure(state=V)
        C.btn_open.configure(state=V)
        C.entry_ean.configure(state=X)
        C._clear_all_slots()

    def _load_existing_files(C):
        """Load images from disk and check FTP copies without blocking GUI."""
        if C.suppress_next_lookup:
            C.suppress_next_lookup = h
            return
        C.logged_counts = h
        F = A.path.join(
            l,
            C.var_name.get().strip().upper(),
            C.var_type.get().strip().upper(),
            C.var_model.get().strip().upper(),
        )
        Y_ = C.var_color1.get().strip().upper()
        Z_ = C.var_color2.get().strip().upper()
        b_ = C.var_color3.get().strip().upper()
        if Y_:
            S_ = [Y_]
            if Z_:
                S_.append(Z_)
            if b_:
                S_.append(b_)
            h_ = g.join(S_)
            F = A.path.join(F, h_)
        I_raw = C.var_extra.get()
        if isinstance(I_raw, dict):
            I_raw = B
        I_ = G(I_raw).strip()
        I_ = I_.replace(a, g)
        if I_ == B:
            I_ = L
        else:
            I_ = I_.upper()
        F = A.path.join(F, I_)
        if I_.upper() == L and not A.path.isdir(F):
            c_ = A.path.join(A.path.dirname(F), L)
            if A.path.isdir(c_):
                try:
                    A.rename(c_, F)
                except E as T:
                    log_error(
                        f"Rename folder NO-LED to NO-LED failed in load_existing_files: {T}"
                    )
        C._clear_all_slots()
        C.original_files = {}
        if not A.path.isdir(F):
            return
        def worker():
            try:
                V_ = [
                    B for B in A.listdir(F) if A.path.isfile(A.path.join(F, B))
                ]
            except E:
                V_ = []
            original_files = {}
            slot_paths = {}
            remote_info = {}
            ean_guess = I
            if V_:
                i_ = V_[0]
                P_ = i_.split(a)
                if P_ and C.var_ean.get().strip() == B:
                    ean_guess = P_[0]
            for W_ in V_:
                d_ = A.path.join(F, W_)
                if not A.path.isfile(d_):
                    continue
                P_ = W_.split(a)
                if Q(P_) < 2:
                    continue
                R_ = P_[1]
                original_files[R_] = W_
                slot_paths[R_] = d_
            ftp_presence = {}
            K_ = C.var_ean.get().strip()
            if K_ and Q(K_) == 13 and K_.isdigit() and K_.upper() != q:
                remote_files = {}
                try:
                    O_ = AB.FTP()
                    O_.connect(D[H][v], D[H][r], timeout=10)
                    O_.login(D[H][N], D[H][M])
                    O_.set_pasv(J)
                    if D[H][m]:
                        O_.cwd(D[H][m])
                    try:
                        e_ = O_.nlst()
                    except AB.error_perm:
                        e_ = []
                    j_ = {A.path.basename(B) for B in e_}
                    for name in j_:
                        if name.startswith(f"{K_}_"):
                            rest = name[len(f"{K_}_"):]
                            label = rest.split(".")[0]
                            remote_files[label] = name
                    for label, fname in remote_files.items():
                        if label not in slot_paths:
                            temp_dir = tempfile.gettempdir()
                            temp_path = A.path.join(temp_dir, fname)
                            try:
                                with x(temp_path, "wb") as fh:
                                    O_.retrbinary(f"RETR {fname}", fh.write)
                                slot_paths[label] = temp_path
                                ftp_presence[label] = fname
                                remote_info[label] = {"filename": fname, "temp_path": temp_path}
                            except E as T:
                                log_error(f"FTP download error for {fname}: {T}")
                        else:
                            ftp_presence[label] = fname
                    O_.quit()
                except E as T:
                    log_error(f"FTP check error for EAN {K_}: {T}")
                if not C.logged_counts:
                    log_info_loc(
                        "found_images_counts",
                        local=Q(original_files),
                        ftp=Q(remote_files),
                    )
                    C.logged_counts = J
            C.after(
                0,
                lambda: finalize(
                    original_files, slot_paths, ftp_presence, remote_info, ean_guess
                ),
            )

        def finalize(original_files, slot_paths, ftp_presence, remote_info, ean_guess):
            if ean_guess and C.var_ean.get().strip() == B:
                C.suppress_next_lookup = J
                C.var_ean.set(ean_guess)
                C.suppress_next_lookup = h
            C.original_files = original_files
            C.ftp_remote_only = remote_info
            C.ftp_presence = ftp_presence
            C.ftp_downloaded_final = set()
            for X_, G_ in A0(C.slots):
                R_ = G_[Aa]
                if R_ in slot_paths:
                    G_[f] = slot_paths[R_]
                    C._update_slot_ui(X_)
                    C._mark_slot(X_, A4)
                else:
                    G_[f] = I
                C._set_icon_status(G_["local_icon"], R_ in original_files)
                C._set_icon_status(G_["ftp_icon"], R_ in ftp_presence)

        threading.Thread(target=worker, daemon=J).start()

    def _on_model_commit(D):
        H = "new"
        o = D.var_name.get().strip()
        p = D.var_type.get().strip()
        e_ = D.var_model.get().strip()
        if not o or not p or not e_:
            return
        if e_.upper() not in D.lists[s]:
            if O.askyesno(
                AJ,
                f"Model '{e_}' nie istnieje na liście. Czy chcesz dodać go do listy?",
            ):
                A6_ = D._open_list_editor(s)
                D.wait_window(A6_)
                D.lists = prepare_excel_lists()
                D.entries = D.lists.get(W, {})
                if W in D.lists:
                    D.lists.pop(W)
                D.combo_model[S] = D.lists[s]
                if e_.upper() not in [A.upper() for A in D.lists[s]]:
                    D.var_model.set(B)
                    return
            else:
                D.var_model.set(B)
                return
        T = A.path.join(l, o.upper(), p.upper(), e_.upper())
        A0_ = []
        if A.path.isdir(T):
            for A1 in A.listdir(T):
                A7 = A.path.join(T, A1)
                if A.path.isdir(A7):
                    A0_.append(A1)
            D.combo_model.configure(style=Z)
        else:
            D.combo_model.configure(style=j)
        r = [A_ for A_ in A0_ if g not in A_]
        A8_ = [A_ for A_ in D.lists[Y] if A_ not in r]
        A9_ = r + A8_
        D._refresh_combobox_list(D.combo_color1, A9_, existing_count=Q(r))
        D.combo_color2[S] = D.lists[Y]
        D.combo_color3[S] = D.lists[Y]
        for AA_ in (D.combo_color1, D.combo_color2, D.combo_color3):
            AA_.configure(state=X)
        D.var_color1.set(B)
        D.var_color2.set(B)
        D.var_color3.set(B)
        D.var_extra.set(B)
        D.var_ean.set(B)
        D.combo_extra.configure(style=j, state=V)
        D.btn_submit.configure(state=V)
        D.btn_open.configure(state=V)
        D._clear_all_slots()
        if not (D.loading_by_ean or D.suppress_scan):
            k_ = []
            if A.path.isdir(T):
                for A2 in A.listdir(T):
                    t_ = A.path.join(T, A2)
                    if A.path.isdir(t_):
                        f_ = A2.split(g)
                        a_ = f_[0] if Q(f_) > 0 else B
                        K__ = f_[1] if Q(f_) > 1 else B
                        M__ = f_[2] if Q(f_) > 2 else B
                        for A3 in A.listdir(t_):
                            AB_ = A.path.join(t_, A3)
                            if A.path.isdir(AB_):
                                u = A3
                                if u.upper() == L or u.upper() == L:
                                    N_ = L
                                else:
                                    N_ = u
                                R_ = q
                                for AC_, b_ in D.entries.items():
                                    if (
                                        b_.get(Ae) == o.upper()
                                        and b_.get(Ad) == p.upper()
                                        and b_.get(AZ) == e_.upper()
                                        and G(b_.get(AY) or B) == a_
                                        and G(b_.get(AX) or B) == K__
                                        and G(b_.get(AW) or B) == M__
                                        and G(b_.get(d) or B) == N_
                                    ):
                                        R_ = AC_
                                        break
                                k_.append((a_, K__, M__, N_, R_))
            if k_:
                if D.model_select_win_open:
                    return
                D.model_select_win_open = J
                P_ = F.Toplevel(D)
                P_.title(SELECT_COMBINATION_TITLE)
                P_.grab_set()
                F.Label(P_, text=SELECT_COMBINATION_PROMPT).pack(pady=5)
                v = C.Frame(P_)
                v.pack(padx=10, fill=z, expand=J)
                m = []
                for AD_ in k_:
                    a_, K__, M__, N_, R_ = AD_
                    w = a_
                    if K__:
                        w += f" / {K__}"
                    if M__:
                        w += f" / {M__}"
                    x = f"{w} - {N_} (EAN: {R_})"
                    m.append(x)
                AE_ = max((Q(A_) for A_ in m), default=0)
                AF_ = max(AE_ + 3, 20)
                i_ = F.Listbox(v, height=5, width=AF_)
                A4_ = C.Scrollbar(v, orient=An, command=i_.yview)
                i_.configure(yscrollcommand=A4_.set)
                A4_.pack(side=AV, fill="y")
                i_.pack(side=Am, fill=z, expand=J)
                for x in m:
                    i_.insert(F.END, x)
                if m:
                    i_.selection_set(0)

                def AG_():
                    A_ = i_.curselection()
                    if not A_:
                        return
                    B_ = A_[0]
                    D.selected_combo = k_[B_]
                    P_.destroy()

                def AH_():
                    D.selected_combo = H
                    P_.destroy()

                n = C.Frame(P_)
                n.pack(pady=5)
                C.Button(n, text=CHOOSE_LABEL, command=AG_).grid(row=0, column=0, padx=5)
                C.Button(n, text=NEW_COMBINATION_LABEL, command=AH_).grid(
                    row=0, column=1, padx=5
                )
                C.Button(n, text=CANCEL_LABEL, command=lambda: P_.destroy()).grid(
                    row=0, column=2, padx=5
                )
                D.selected_combo = I
                D.wait_window(P_)
                D.model_select_win_open = h
                y_ = Aj(D, "selected_combo", I)
                D.selected_combo = I
                if y_ and y_ != H:
                    a_, K__, M__, N_, R_ = y_
                    D.var_color1.set(a_)
                    D.var_color2.set(K__)
                    D.var_color3.set(M__)
                    AI_ = g.join([A_ for A_ in (a_, K__, M__) if A_])
                    c_ = A.path.join(T, AI_)
                    H_ = []
                    if A.path.isdir(c_):
                        H_ = [
                            B for B in A.listdir(c_) if A.path.isdir(A.path.join(c_, B))
                        ]
                        D.combo_color1.configure(style=Z)
                        if K__:
                            D.combo_color2.configure(style=Z)
                        if M__:
                            D.combo_color3.configure(style=Z)
                    else:
                        D.combo_color1.configure(style=j)
                        if K__:
                            D.combo_color2.configure(style=j)
                        if M__:
                            D.combo_color3.configure(style=j)
                    AK_ = [A_ for A_ in D.lists[d] if A_ not in H_]
                    if L in H_ and L not in H_:
                        try:
                            A.rename(A.path.join(c_, L), A.path.join(c_, L))
                        except E as AL_:
                            log_error(f"Rename folder NO-LED to NO-LED failed: {AL_}")
                        H_ = [
                            B for B in A.listdir(c_) if A.path.isdir(A.path.join(c_, B))
                        ]
                        if L in H_:
                            H_[H_.index(L)] = L
                    D._refresh_combobox_list(
                        D.combo_extra, H_ + AK_, existing_count=Q(H_)
                    )
                    D.combo_extra.configure(state=X)
                    if N_ == L:
                        D.var_extra.set(B)
                    else:
                        D.var_extra.set(N_)
                    if R_ and G(R_).upper() != q:
                        D.var_ean.set(R_)
                    else:
                        D.var_ean.set(q)
                    D.combo_extra.configure(
                        style=Z if N_ in H_ or N_ == L and L in H_ else j
                    )
                    D.combo_model.configure(style=Z)
                    D.combo_color1.configure(style=Z)
                    if K__:
                        D.combo_color2.configure(style=Z)
                    if M__:
                        D.combo_color3.configure(style=Z)
                    D._load_existing_files()
                    D.btn_submit.configure(state=X)
                    D.btn_open.configure(state=X)

    def _on_key_release(C, event):
        J_ = event
        A_ = J_.widget
        if J_.keysym in ("Up", "Down", "Left", "Right"):
            return
        D_ = I
        if A_ == C.combo_name:
            D_ = n
        elif A_ == C.combo_type:
            D_ = t
        elif A_ == C.combo_model:
            D_ = s
        elif A_ in (C.combo_color1, C.combo_color2, C.combo_color3):
            D_ = Y
        elif A_ == C.combo_extra:
            D_ = d
        else:
            return
        E_ = A_.get()
        if E_ == B:
            A_[S] = C.lists[D_]
            return
        H_ = [A for A in C.lists[D_] if A and A.lower().startswith(E_.lower())]
        if H_:
            H_.sort(key=G.lower)
            A_[S] = H_
            if J_.keysym not in ("BackSpace", "Delete"):
                K_ = H_[0]
                if E_.lower() != K_.lower():
                    A_.set(K_)
                    A_.icursor(Q(E_))
                    A_.selection_range(Q(E_), F.END)
        else:
            A_[S] = []

    def _on_color_commit(C):
        M_ = C.var_name.get().strip()
        N_ = C.var_type.get().strip()
        H_ = C.var_color1.get().strip()
        F_ = C.var_color2.get().strip()
        G_ = C.var_color3.get().strip()
        if C.var_ean.get().strip():
            C.var_ean.set(B)
        if not M_ or not N_ or not H_:
            return
        J_ = [A for A in (H_, F_, G_) if A and A.upper() not in C.lists[Y]]
        if J_:
            P_ = AI.join(J_)
            R_ = (
                f"Kolor '{J_[0]}' nie istnieje na liście. Czy dodać nowy wpis?"
                if Q(J_) == 1
                else f"Kolory '{P_}' nie istnieją na liście. Czy dodać nowe wpisy?"
            )
            if O.askyesno(AJ, R_):
                T = C._open_list_editor(Y)
                C.wait_window(T)
                C.lists = prepare_excel_lists()
                C.entries = C.lists.get(W, {})
                if W in C.lists:
                    C.lists.pop(W)
                C.combo_color1[S] = C.lists[Y]
                C.combo_color2[S] = C.lists[Y]
                C.combo_color3[S] = C.lists[Y]
                if H_.upper() not in [A.upper() for A in C.lists[Y]]:
                    C.var_color1.set(B)
                    return
                if F_ and F_.upper() not in [A.upper() for A in C.lists[Y]]:
                    C.var_color2.set(B)
                if G_ and G_.upper() not in [A.upper() for A in C.lists[Y]]:
                    C.var_color3.set(B)
            else:
                if H_.upper() not in [A.upper() for A in C.lists[Y]]:
                    C.var_color1.set(B)
                    return
                if F_ and F_.upper() not in [A.upper() for A in C.lists[Y]]:
                    C.var_color2.set(B)
                if G_ and G_.upper() not in [A.upper() for A in C.lists[Y]]:
                    C.var_color3.set(B)
        H_ = C.var_color1.get().strip()
        if not H_:
            return
        K_ = [H_]
        if F_:
            K_.append(F_)
        if G_:
            K_.append(G_)
        V_ = g.join(K_)
        I_ = A.path.join(
            l, M_.upper(), N_.upper(), C.var_model.get().strip().upper(), V_
        )
        D_ = []
        if A.path.isdir(I_):
            D_ = [B for B in A.listdir(I_) if A.path.isdir(A.path.join(I_, B))]
            if L in D_ and L not in D_:
                try:
                    A.rename(A.path.join(I_, L), A.path.join(I_, L))
                except E as a_:
                    log_error(f"Rename folder NO-LED to NO-LED failed: {a_}")
                D_ = [B for B in A.listdir(I_) if A.path.isdir(A.path.join(I_, B))]
            if L in D_:
                D_[D_.index(L)] = L
            C.combo_color1.configure(style=Z)
            if F_:
                C.combo_color2.configure(style=Z)
            if G_:
                C.combo_color3.configure(style=Z)
        else:
            C.combo_color1.configure(style=j)
            if F_:
                C.combo_color2.configure(style=j)
            if G_:
                C.combo_color3.configure(style=j)
        b_ = [A for A in C.lists[d] if A not in D_]
        C._refresh_combobox_list(C.combo_extra, D_ + b_, existing_count=Q(D_))
        C.combo_extra.configure(state=X)
        C.entry_ean.configure(state=X)
        C.btn_submit.configure(state=X)
        C.btn_open.configure(state=X)
        extra_raw = C.var_extra.get()
        C.var_extra.set(G(extra_raw).strip())
        if not C.suppress_scan:
            C._load_existing_files()

    def _on_extra_commit(C):
        D_ = C.var_extra.get().strip()
        G_ = C.var_name.get().strip()
        H_ = C.var_type.get().strip()
        I_ = C.var_model.get().strip()
        F_ = C.var_color1.get().strip()
        J_ = C.var_color2.get().strip()
        K_ = C.var_color3.get().strip()
        if D_ == B:
            C.combo_extra.configure(style=j)
        else:
            if D_.upper() not in [A.upper() for A in C.lists[d]]:
                if O.askyesno(
                    AJ,
                    VALUE_NOT_EXISTS_QUESTION.format(value=D_),
                ):
                    M_ = C._open_list_editor(d)
                    C.wait_window(M_)
                    C.lists = prepare_excel_lists()
                    C.entries = C.lists.get(W, {})
                    if W in C.lists:
                        C.lists.pop(W)
                    C.combo_extra[S] = C.lists[d]
                    if D_.upper() not in [A.upper() for A in C.lists[d]]:
                        C.var_extra.set(B)
                        D_ = B
                else:
                    C.var_extra.set(B)
                    D_ = B
                    C.combo_extra.configure(style=j)
                    return
            E_ = A.path.join(
                l, G_.upper(), H_.upper(), I_.upper(), F_.upper() if F_ else B
            )
            if J_:
                E_ = A.path.join(E_, J_.upper())
                if K_:
                    E_ = A.path.join(E_, K_.upper())
            N_ = D_.strip().replace(a, g).upper() if D_ else L
            E_ = A.path.join(E_, N_)
            if A.path.isdir(E_):
                C.combo_extra.configure(style=Z)
            else:
                C.combo_extra.configure(style=j)
        if G_ and H_ and I_ and F_ and not C.suppress_scan:
            C._load_existing_files()

    def _select_file(A, idx):
        if A.is_processing:
            O.showwarning(OPERATION_TITLE, PROCESSING_MSG)
            return
        if not (
            A.var_name.get().strip()
            and A.var_type.get().strip()
            and A.var_model.get().strip()
            and A.var_color1.get().strip()
        ):
            O.showwarning(INCOMPLETE_DATA_MSG, MISSING_FIELDS_MSG)
            return
        C_ = [
            ("Obrazy/PDF/DOC", "*.jpg *.jpeg *.png *.pdf *.doc *.docx"),
            ("Wszystkie pliki", "*.*"),
        ]
        B_ = BT.askopenfilename(title=SELECT_FILE_TITLE, filetypes=C_)
        if B_:
            A._add_file_to_slot(idx, B_)

    def _on_drop(C, event, idx):
        if C.is_processing:
            return
        if not (
            C.var_name.get().strip()
            and C.var_type.get().strip()
            and C.var_model.get().strip()
            and C.var_color1.get().strip()
        ):
            O.showwarning(INCOMPLETE_DATA_MSG, MISSING_FIELDS_MSG)
            return
        G_ = C.tk.splitlist(event.data)
        if G_:
            C._add_file_to_slot(idx, G_[0])
        if C.dragging_idx is not I:
            D_ = C.dragging_idx
            if D_ != idx:
                H_ = h
                E_ = C.slots[D_][f]
                if E_:
                    if D_ in C.pending_additions:
                        C.pending_additions.pop(D_, I)
                        H_ = J
                    elif E_.startswith(l) and A.path.isfile(E_):
                        C.pending_deletions[D_] = E_
                    C.slots[D_][f] = I
                    F_ = C.slots[D_]
                    F_[y].configure(image=B, text=NO_FILE_LABEL)
                    F_[y].image = I
                    F_[A7].place_forget()
                    if H_:
                        C._mark_slot(D_, I)
                    else:
                        C._mark_slot(D_, AR)
                    C.focus_force()
            C.dragging_idx = I

    def _add_file_to_slot(B, idx, src_path):
        E_ = src_path
        C_ = idx
        D_ = B.slots[C_][f]
        if D_:
            if C_ in B.pending_additions:
                B.pending_additions.pop(C_, I)
            elif D_.startswith(l) and A.path.isfile(D_):
                B.pending_deletions[C_] = D_
        F_ = B.var_ean.get().strip()
        if not F_:
            F_ = q
        B.pending_additions[C_] = E_
        B.slots[C_][f] = E_
        B._update_slot_ui(C_)
        B.slots[C_][A7].place(x=0, y=0)
        B._mark_slot(C_, AR)
        B._set_icon_status(B.slots[C_]["local_icon"], J)

    def _update_slot_ui(J, idx):
        D_ = J.slots[idx]
        F_ = D_[f]
        C_ = D_[y]
        K_ = D_[A7]
        if not F_:
            return
        try:
            G_ = AA.open(F_)
            G_.thumbnail((100, 100), AA.LANCZOS)
            H_ = ImageTk.PhotoImage(G_)
            C_.configure(image=H_, text=B)
            C_.image = H_
        except E:
            C_.configure(text=A.path.basename(F_), image=B)
            C_.image = I
        K_.place(x=0, y=0)

    def _remove_file(C, idx):
        if C.is_processing:
            O.showwarning(OPERATION_TITLE, PROCESSING_MSG)
            return
        D_ = idx
        E_ = C.slots[D_]
        F_ = E_[f]
        if F_:
            if not O.askyesno(
                "Usuń plik", f"Czy na pewno usunąć plik {A.path.basename(F_)}?"
            ):
                return
            G_ = h
            if D_ in C.pending_additions:
                C.pending_additions.pop(D_, I)
                G_ = J
            elif F_.startswith(l) and A.path.isfile(F_):
                C.pending_deletions[D_] = F_
            elif not F_.startswith(l):
                label = E_[Aa]
                remote_name = I
                info = C.ftp_remote_only.pop(label, I)
                if info:
                    remote_name = info.get("filename")
                elif label in C.ftp_presence:
                    remote_name = C.ftp_presence.get(label)
                if remote_name:
                    C.pending_ftp_deletions[D_] = remote_name
            E_[f] = I
            E_[y].configure(image=B, text=NO_FILE_LABEL)
            E_[y].image = I
            E_[A7].place_forget()
            C._set_icon_status(E_["local_icon"], h)
            if G_:
                C._mark_slot(D_, I)
            else:
                C._mark_slot(D_, AR)
            C.focus_force()

    def _clear_all_slots(C):
        C.pending_additions.clear()
        C.pending_deletions.clear()
        C.pending_ftp_deletions.clear()
        for A_ in C.slots:
            A_[f] = I
            A_[y].configure(image=B, text=NO_FILE_LABEL)
            A_[y].image = I
            A_[A7].place_forget()
            A_["local_icon"].place_forget()
            A_["local_icon"].delete("slash")
            A_["ftp_icon"].place_forget()
            A_["ftp_icon"].delete("slash")
            if AS in A_:
                A_[AS].configure(
                    highlightthickness=0, highlightbackground=A8, highlightcolor=A8
                )

    def _open_list_editor(E, focus_sheet=I):
        H_ = F.Toplevel(E)
        H_.title(EDIT_LISTS_LABEL)
        H_.grab_set()
        I_ = C.Notebook(H_)
        I_.pack(expand=J, fill=z, padx=5, pady=5)
        M_ = {}
        P_ = [(n, "Nazwy"), (t, "Typy"), (s, "Modele"), (Y, "Kolory"), (d, "Dodatki")]
        N_ = 0
        for R_, (A_, S_) in A0(P_):
            B_ = C.Frame(I_)
            I_.add(B_, text=S_)
            M_[A_] = B_
            if focus_sheet == A_:
                N_ = R_
        I_.select(N_)
        K_ = 0
        for T in (n, t, s, Y, d):
            for G_ in E.lists[T]:
                if G_ and Q(G_) > K_:
                    K_ = Q(G_)
        U = max(K_ + 3, 20)
        for A_, B_ in M_.items():
            V_ = E.lists[A_]
            D_ = F.Listbox(B_, height=5, width=U)
            O_ = C.Scrollbar(B_, orient=An, command=D_.yview)
            D_.configure(yscrollcommand=O_.set)
            L_ = C.Frame(B_)
            L_.pack(side=AV, fill="y", padx=5, pady=5)
            O_.pack(side=AV, fill="y", pady=5)
            D_.pack(side=Am, fill=z, expand=J, padx=5, pady=5)
            for G_ in V_:
                D_.insert(F.END, G_)
            C.Button(
                L_, text="Dodaj", command=lambda k=A_, l=D_: E._add_list_item(k, l)
            ).pack(fill="x", pady=2)
            C.Button(
                L_, text="Usuń", command=lambda k=A_, l=D_: E._remove_list_item(k, l)
            ).pack(fill="x", pady=2)
        return H_

    def _add_list_item(C, key, listbox):
        B_ = key
        D_ = BI.askstring("Dodaj", f"Nowa wartość do listy {B_}:")
        if D_:
            add_to_list(AE[B_], D_)
            if D_.strip().upper() not in [A.upper() for A in C.lists[B_]]:
                C.lists[B_] = C.lists[B_] + [
                    D_.strip().upper() if B_ != d else D_.strip().replace(a, g).upper()
                ]
            listbox.insert(
                F.END,
                D_.strip().upper() if B_ != d else D_.strip().replace(a, g).upper(),
            )
            if B_ == n:
                C.combo_name[S] = C.lists[B_]
            elif B_ == t:
                C.combo_type[S] = C.lists[B_]
            elif B_ == s:
                C.combo_model[S] = C.lists[B_]
            elif B_ == Y:
                C.combo_color1[S] = C.lists[B_]
                C.combo_color2[S] = C.lists[B_]
                C.combo_color3[S] = C.lists[B_]
            elif B_ == d:
                C.combo_extra[S] = C.lists[B_]

    def _remove_list_item(A, key, listbox):
        D_ = listbox
        B_ = key
        E_ = D_.curselection()
        if not E_:
            return
        F_ = E_[0]
        C_ = D_.get(F_)
        if O.askyesno("Usuń", f"Czy usunąć '{C_}' z listy {B_}?"):
            remove_from_list(AE[B_], C_)
            if C_ in A.lists[B_] or C_.upper() in [A.upper() for A in A.lists[B_]]:
                A.lists[B_] = [
                    A_ for A_ in A.lists[B_] if A_.upper() != C_.strip().upper()
                ]
            D_.delete(F_)
            if B_ == n:
                A.combo_name[S] = A.lists[B_]
            elif B_ == t:
                A.combo_type[S] = A.lists[B_]
            elif B_ == s:
                A.combo_model[S] = A.lists[B_]
            elif B_ == Y:
                A.combo_color1[S] = A.lists[B_]
                A.combo_color2[S] = A.lists[B_]
                A.combo_color3[S] = A.lists[B_]
            elif B_ == d:
                A.combo_extra[S] = A.lists[B_]

    def _on_submit(C):
        A2 = "was_existing"
        t = "inter_set"
        s = "del_set"
        p = "add_set"
        o = "pending_del_leftover"
        n = "pending_add_leftover"
        k = "ftp_skipped"
        j = "sql_rows"
        d = "sql_queries"
        c = "ftp_time"
        b = "ftp_deleted"
        Z = "ftp_sent"
        Y = "ftp_error_msg"
        S = "sql_time"
        P = "sql_error_msg"
        K = "error_set"
        if not (
            C.var_name.get().strip()
            and C.var_type.get().strip()
            and C.var_model.get().strip()
            and C.var_color1.get().strip()
        ):
            O.showwarning(
                NO_DATA_MSG,
                FILL_REQUIRED_BEFORE_SUBMIT_MSG,
            )
            return
        if C.var_extra.get().strip() == B:
            C.var_extra.set(L)
        if not C.var_ean.get().strip():
            Ai_ = BI.askstring(
                EAN_PROMPT_TITLE,
                EAN_MISSING_PROMPT,
            )
            if Ai_ is I or Ai_.strip() == B:
                Ai_ = q
            C.var_ean.set(Ai_.strip())
        AE_ = C.var_name.get().strip()
        AF_ = C.var_type.get().strip()
        AG_ = C.var_model.get().strip()
        AH_ = C.var_color1.get().strip()
        p_ = C.var_color2.get().strip()
        s_ = C.var_color3.get().strip()
        b_ = C.var_extra.get().strip()
        if b_ == B or b_.upper() in [L, L]:
            b_ = L
        else:
            b_ = b_.replace(a, g).upper()
        K_ = C.var_ean.get().strip()
        BY_ = K_.upper() != q and K_ in C.entries
        BZ_ = save_ean_entry(
            K_, AE_, AF_, AG_, AH_, p_ or B, s_ or B, b_ if b_ != B else L
        )
        if BZ_ is h:
            return
        else:
            try:
                BC_ = prepare_excel_lists()
                if W in BC_:
                    C.entries = BC_[W]
            except E as R:
                log_error(f"Failed to reload entries after saving: {R}")
        C.is_processing = J
        C.btn_submit.configure(state=V)
        C.btn_open.configure(state=V)
        for widget in [
            C.combo_name,
            C.combo_type,
            C.combo_model,
            C.combo_color1,
            C.combo_color2,
            C.combo_color3,
            C.combo_extra,
            C.entry_ean,
        ]:
            try:
                widget.configure(state=Ak)
            except:
                pass
        C.ui_log.configure(state=Az)
        C.ui_log.insert(F.END, PROCESSING_UI_MSG + "\n")
        C.ui_log.configure(state=Ak)
        result_data = {}

        def heavy_work():
            A3 = "rowcount"
            X = "optimize"
            W = "quality"
            V = ".png"
            O = ".jpeg"
            F = ".jpg"
            result_data[Z] = 0
            result_data[b] = 0
            result_data[c] = 0
            result_data[d] = 0
            result_data[j] = 0
            result_data[S] = 0
            result_data[K] = set()
            result_data[Y] = B
            result_data[k] = Ay
            result_data[P] = B
            try:
                i_ = A.path.join(l, AE_.upper(), AF_.upper(), AG_.upper())
                Av_ = [AH_.upper()]
                if p_:
                    Av_.append(p_.upper())
                if s_:
                    Av_.append(s_.upper())
                BX_ = g.join(Av_)
                i_ = A.path.join(i_, BX_, b_ if b_ != B else L)
                A.makedirs(i_, exist_ok=J)
                BM_ = []
                files_to_upload = []
                try:
                    if A.path.exists(AN):
                        Af.rmtree(AN)
                    A.makedirs(AN, exist_ok=J)
                except E as R:
                    log_error_loc("backup_folder_failed", error=R)
                backed_up = []
                for T in set(C.pending_deletions.values()):
                    if T and A.path.isfile(T):
                        try:
                            Af.copy2(T, A.path.join(AN, A.path.basename(T)))
                            backed_up.append(A.path.basename(T))
                        except E as R:
                            log_error_loc(
                                "backup_file_failed",
                                file=A.path.basename(T),
                                error=R,
                            )
                if backed_up:
                    log_info_loc(
                        "backup_files_done", files=AI.join(backed_up)
                    )
                if C.ftp_remote_only:
                    for label, info in C.ftp_remote_only.items():
                        for idx, slot in A0(C.slots):
                            if slot[Aa] == label:
                                Az_ = Aw[idx][0]
                                Be_ = label_category(Aw[idx][1])
                                P_ = [
                                    K_ if K_ else q,
                                    Az_,
                                    Be_,
                                    AE_.upper(),
                                    AF_.upper(),
                                    AG_.upper(),
                                    AH_.upper(),
                                ]
                                if p_:
                                    P_.append(p_.upper())
                                if s_:
                                    P_.append(s_.upper())
                                P_.append(b_ if b_ != B else L)
                                ext = A.path.splitext(info["filename"])[1]
                                c_ = a.join(P_) + ext
                                dest = A.path.join(i_, c_)
                                try:
                                    Af.copy2(info["temp_path"], dest)
                                    log_info_loc(
                                        "ftp_file_downloaded",
                                        file=info["filename"],
                                        temp=c_,
                                    )
                                    files_to_upload.append(c_)
                                    C.slots[idx][f] = dest
                                    C.ftp_downloaded_final.add(c_)
                                except E as R:
                                    log_error_loc(
                                        "file_save_error",
                                        file=info["filename"],
                                        error=R,
                                    )
                                break
                    C.ftp_remote_only = {}
                AJ_ = set(C.pending_additions.keys())
                AL_ = set(C.pending_deletions.keys())
                AM_ = AJ_ & AL_
                for F_ in list(AM_):
                    A8_ = C.pending_additions.get(F_)
                    Ay_ = C.pending_deletions.get(F_)
                    if A8_ and Ay_:
                        try:
                            BD_ = A.path.samefile(A8_, Ay_)
                        except E:
                            BD_ = A.path.normcase(
                                A.path.normpath(A8_)
                            ) == A.path.normcase(A.path.normpath(Ay_))
                        if BD_:
                            C.pending_additions.pop(F_, I)
                            C.pending_deletions.pop(F_, I)
                AJ_ = set(C.pending_additions.keys())
                AL_ = set(C.pending_deletions.keys())
                AM_ = AJ_ & AL_
                BE_ = {}
                for F_, src_path in list(C.pending_additions.items()):
                    if F_ not in C.pending_deletions and C.slots[F_].get(B0) != AR:
                        C.pending_additions.pop(F_, I)
                        continue
                    if not A.path.isfile(src_path):
                        C.pending_additions.pop(F_, I)
                        continue
                    Az_ = Aw[F_][0]
                    Be_ = label_category(Aw[F_][1])
                    P_ = [
                        K_ if K_ else q,
                        Az_,
                        Be_,
                        AE_.upper(),
                        AF_.upper(),
                        AG_.upper(),
                        AH_.upper(),
                    ]
                    if p_:
                        P_.append(p_.upper())
                    if s_:
                        P_.append(s_.upper())
                    P_.append(b_ if b_ != B else L)
                    BH_ = A.path.splitext(src_path)[1]
                    c_ = a.join(P_) + BH_
                    if F_ in C.pending_ftp_deletions and C.pending_ftp_deletions[F_] == c_:
                        C.pending_ftp_deletions.pop(F_, I)
                    S_ = A.path.join(i_, c_)
                    try:
                        if F_ in C.pending_deletions:
                            old_path = C.pending_deletions.get(F_)
                            try:
                                same_target = A.path.samefile(old_path, S_)
                            except E:
                                same_target = A.path.normcase(
                                    A.path.normpath(old_path)
                                ) == A.path.normcase(A.path.normpath(S_))
                            if same_target:
                                C.pending_deletions.pop(F_, I)
                                try:
                                    if A.path.exists(old_path):
                                        A.remove(old_path)
                                        log_info_loc(
                                            "deleted_file_before_add",
                                            file=A.path.basename(old_path),
                                        )
                                except E as z:
                                    log_error_loc(
                                        "remove_old_file_failed",
                                        file=A.path.basename(old_path),
                                        error=z,
                                    )
                            elif A.path.exists(S_):
                                try:
                                    A.remove(S_)
                                except E as z:
                                    log_error_loc(
                                        "remove_file_before_overwrite_failed",
                                        file=A.path.basename(S_),
                                        error=z,
                                    )
                        elif A.path.exists(S_):
                            try:
                                A.remove(S_)
                            except E as z:
                                log_error_loc(
                                    "remove_file_before_overwrite_failed",
                                    file=A.path.basename(S_),
                                    error=z,
                                )
                        ext_lower = BH_.lower()
                        if ext_lower in [F, O, V, ".bmp", ".gif"]:
                            A1 = AA.open(src_path)
                            if C.opt_resize.get():
                                max_dim = C.resize_max_dim.get() or 2000
                                A1.thumbnail((max_dim, max_dim), AA.LANCZOS)
                            save_params = {}
                            if ext_lower in [F, O]:
                                quality = 95
                                if C.opt_compress.get():
                                    quality = max(
                                        1, min(100, C.compress_quality.get() or 85)
                                    )
                                save_params[W] = quality
                                save_params[X] = J
                            if ext_lower == V:
                                save_params[X] = J
                            A1.save(S_, **save_params)
                            if C.opt_maxsize.get():
                                max_bytes = (C.max_file_kb.get() or 0) * 1024
                                if max_bytes > 0:
                                    if A.path.getsize(S_) > max_bytes and ext_lower in [
                                        F,
                                        O,
                                    ]:
                                        try:
                                            quality = save_params.get(W, 95)
                                            while (
                                                quality > 10
                                                and A.path.getsize(S_) > max_bytes
                                            ):
                                                quality -= 5
                                                A1.save(S_, quality=quality, optimize=J)
                                        except E as R:
                                            log_error_loc(
                                                "file_resize_error",
                                                file=c_,
                                                error=R,
                                            )
                            log_info_loc("image_added_modified", file=c_)
                        elif ext_lower in [".tif", ".tiff"]:
                            if C.opt_convert_tif.get():
                                target_fmt = C.tif_target_format.get().upper()
                                if target_fmt in ["JPG", "JPEG"]:
                                    t_ext = F
                                elif target_fmt == "PNG":
                                    t_ext = V
                                elif target_fmt == "BMP":
                                    t_ext = ".bmp"
                                elif target_fmt == "GIF":
                                    t_ext = ".gif"
                                else:
                                    t_ext = "." + target_fmt.lower()
                                c_ = a.join(P_) + t_ext
                                S_ = A.path.join(i_, c_)
                                if A.path.exists(S_):
                                    try:
                                        A.remove(S_)
                                    except E as z:
                                        log_error_loc(
                                            "remove_file_before_overwrite_failed",
                                            file=A.path.basename(S_),
                                            error=z,
                                        )
                                A1 = AA.open(src_path)
                                if C.opt_resize.get():
                                    max_dim = C.resize_max_dim.get() or 2000
                                    A1.thumbnail((max_dim, max_dim), AA.LANCZOS)
                                save_params = {}
                                if t_ext in [F, O]:
                                    quality = 95
                                    if C.opt_compress.get():
                                        quality = max(
                                            1, min(100, C.compress_quality.get() or 85)
                                        )
                                    save_params[W] = quality
                                    save_params[X] = J
                                if t_ext == V:
                                    save_params[X] = J
                                A1.save(S_, **save_params)
                                if C.opt_maxsize.get():
                                    max_bytes = (C.max_file_kb.get() or 0) * 1024
                                    if max_bytes > 0 and t_ext in [F, O]:
                                        try:
                                            quality = save_params.get(W, 95)
                                            while (
                                                quality > 10
                                                and A.path.getsize(S_) > max_bytes
                                            ):
                                                quality -= 5
                                                A1.save(S_, quality=quality, optimize=J)
                                        except E as R:
                                            log_error_loc(
                                                "file_resize_error",
                                                file=c_,
                                                error=R,
                                            )
                                log_info_loc("image_added_modified", file=c_)
                            else:
                                Af.copy2(src_path, S_)
                                log_info_loc("file_added_modified", file=c_)
                        else:
                            Af.copy2(src_path, S_)
                            log_info_loc("file_added_modified", file=c_)
                        files_to_upload.append(c_)
                        C.slots[F_][f] = S_
                    except E as y:
                        log_error_loc(
                            "file_copy_failed",
                            file=A.path.basename(src_path),
                            error=y,
                        )
                        result_data[K].add(F_)
                        BE_[F_] = src_path
                        continue
                if K_ and Q(K_) == 13 and K_.isdigit():
                    try:
                        file_list = A.listdir(i_)
                    except E:
                        file_list = []
                    remove_candidates = {
                        A.path.basename(B) for B in C.pending_deletions.values()
                    }
                    for X_ in file_list:
                        path = A.path.join(i_, X_)
                        if not A.path.isfile(path):
                            continue
                        if X_ in remove_candidates:
                            continue
                        P_ = X_.split(a)
                        ean_prefix = P_[0] if P_ else B
                        if ean_prefix.upper() != K_.upper():
                            new_name = K_ + a + a.join(P_[1:]) if Q(P_) > 1 else K_
                            new_path = A.path.join(i_, new_name)
                            try:
                                if A.path.exists(new_path):
                                    A.remove(new_path)
                                A.rename(path, new_path)
                                log_info_loc(
                                    "file_renamed", old=X_, new=new_name
                                )
                                for F_, d_ in A0(C.slots):
                                    if d_[f] and A.path.basename(d_[f]) == X_:
                                        C.slots[F_][f] = new_path
                                        break
                                if X_ in files_to_upload:
                                    Bh_ = files_to_upload.index(X_)
                                    files_to_upload[Bh_] = new_name
                            except E as y:
                                log_error_loc(
                                    "file_rename_error", ean=K_, error=y
                                )
                                for i, d_ in A0(C.slots):
                                    if d_[f] and A.path.basename(d_[f]) == X_:
                                        result_data[K].add(i)
                                        break
                for idx, slot in A0(C.slots):
                    path = slot[f]
                    if (
                        path
                        and A.path.isfile(path)
                        and idx not in C.pending_deletions
                        and slot[Aa] not in C.ftp_presence
                    ):
                        fname = A.path.basename(path)
                        if fname not in files_to_upload:
                            files_to_upload.append(fname)
                        C.pending_additions.setdefault(idx, path)
                Am_ = {}
                for F_, T in list(C.pending_deletions.items()):
                    if F_ in result_data[K]:
                        Am_[F_] = T
                        continue
                    conflict_error = h
                    for Bh in result_data[K]:
                        if C.pending_additions.get(Bh) == T:
                            conflict_error = J
                            break
                    if conflict_error:
                        Am_[F_] = T
                        continue
                    try:
                        if A.path.isfile(T):
                            A.remove(T)
                            log_info_loc(
                                "file_deleted", file=A.path.basename(T)
                            )
                            BO_ = A.path.basename(T)
                            P_ = BO_.split(a)
                            if Q(P_) >= 2:
                                An_ = P_[0]
                                Bi = P_[1]
                                Bj = A.path.splitext(BO_)[1]
                                if An_ and Q(An_) == 13 and An_.isdigit():
                                    BM_.append(f"{An_}_{Bi}{Bj}")
                    except E as y:
                        log_error_loc(
                            "file_delete_failed",
                            file=A.path.basename(T),
                            error=y,
                        )
                        result_data[K].add(F_)
                        Am_[F_] = T
                for Cz_ in C.pending_ftp_deletions.values():
                    if Cz_:
                        BM_.append(Cz_)
                result_data[n] = BE_
                result_data[o] = Am_
                add_set = set(C.pending_additions.keys())
                del_set = set(C.pending_deletions.keys())
                inter_set = add_set & del_set
                result_data[p] = add_set
                result_data[s] = del_set
                result_data[t] = inter_set
                A__ = Ay
                Y_ = B
                BQ = 0
                BR_ = 0
                Bk = Ag.perf_counter()
                if not result_data[K]:
                    if not (K_ and Q(K_) == 13 and K_.isdigit()):
                        A__ = J
                    elif not D.get(ft, J):
                        log_info_loc("ftp_upload_skipped_settings")
                    else:
                        ftp = AB.FTP()
                        try:
                            ftp.connect(D[H][v], D[H][r], timeout=10)
                            ftp.login(D[H][N], D[H][M])
                            ftp.set_pasv(J)
                            if D[H][m]:
                                ftp.cwd(D[H][m])
                        except AB.error_perm as R:
                            AT = G(R)
                            if "530" in AT or LOGIN_INCORRECT_MSG in AT:
                                Y_ = LOGIN_DATA_ERROR_MSG
                            elif As in AT or NO_SUCH_FILE_MSG in AT:
                                Y_ = PATH_NOT_FOUND_MSG
                            else:
                                Y_ = f"Błąd FTP: {AT}"
                        except (
                            BK.gaierror,
                            CONNECTION_REFUSED_ERROR,
                            TIMEOUT_ERROR,
                            Au,
                        ) as R:
                            Y_ = NETWORK_ERROR_MSG
                        except E as R:
                            Y_ = f"Inny błąd: {R}"
                        else:
                            try:
                                files_local = [
                                    B
                                    for B in files_to_upload
                                    if A.path.isfile(A.path.join(i_, B))
                                ]
                                ftp_error = h
                                for X_ in files_local:
                                    if X_ in C.ftp_downloaded_final:
                                        log_info_loc(
                                            "ftp_upload_skipped_downloaded", file=X_
                                        )
                                        continue
                                    P_ = X_.split(a)
                                    Ao_ = P_[0] if P_ else B
                                    if not (Ao_ and Q(Ao_) == 13 and Ao_.isdigit()):
                                        continue
                                    Bl = P_[1] if Q(P_) > 1 else B
                                    Bm = A.path.splitext(X_)[1]
                                    BT = f"{Ao_}_{Bl}{Bm}"
                                    Bn = A.path.join(i_, X_)
                                    try:
                                        with x(Bn, "rb") as Bo:
                                            ftp.storbinary(f"STOR {BT}", Bo)
                                            BQ += 1
                                            log_info_loc(
                                                "ftp_file_uploaded", file=X_, target=BT
                                            )
                                    except E as AU:
                                        Y_ = f"Błąd wysyłania pliku {X_}: {AU}"
                                        log_error_loc(
                                            "ftp_upload_error_file",
                                            file=X_,
                                            error=AU,
                                        )
                                        ftp_error = J
                                        break
                                if not ftp_error:
                                    Ap = []
                                    for AV_ in BM_:
                                        try:
                                            ftp.delete(AV_)
                                            BR_ += 1
                                            log_info_loc(
                                                "ftp_file_deleted", file=AV_
                                            )
                                        except E as AU:
                                            Bp = G(AU)
                                            if As in Bp:
                                                log_info_loc(
                                                    "ftp_file_missing_no_delete",
                                                    file=AV_,
                                                )
                                            else:
                                                Ap.append(AV_)
                                                log_error_loc(
                                                    "ftp_delete_error",
                                                    file=AV_,
                                                    error=AU,
                                                )
                                    if Ap:
                                        if not Y_:
                                            Y_ = f"Nie udało się usunąć niektórych plików na FTP: {AI.join(Ap)}"
                                        else:
                                            Y_ += f". Nie udało się usunąć plików: {AI.join(Ap)}"
                            finally:
                                try:
                                    ftp.quit()
                                except E:
                                    pass
                result_data[Y] = Y_
                result_data[k] = A__
                result_data[Z] = BQ
                result_data[b] = BR_
                Bq = int((Ag.perf_counter() - Bk) * 1000)
                result_data[c] = Bq
                AW_ = B
                Aq_ = 0
                CANCEL_LABEL = 0
                INCOMPLETE_DATA_MSG = 0
                if D.get(u, J) and K_ and len(K_) == 13 and K_.isdigit():
                    Br = Ag.perf_counter()
                    try:
                        conn = connect_db()
                        cur = conn.cursor()
                        for d_ in C.slots:
                            Az_ = d_[Aa]
                            B3_ = d_["label"]
                            if d_[f]:
                                Bs = A.path.basename(d_[f])
                                ext = A.path.splitext(Bs)[1].lower()
                                short_name = f"{K_}_{Az_}{ext}"
                                try:
                                    AX_ = D.get(w, SQL_UPDATE_TEMPLATE)
                                    AC_ = AX_.format(
                                        col=B3_, filename=short_name, ean=K_
                                    )
                                except E as R:
                                    raise E(f"Błąd formatowania zapytania SQL: {R}")
                                cur.execute(AC_)
                                Aq_ += 1
                                if Aj(cur, A3, -1) >= 0:
                                    CANCEL_LABEL += cur.rowcount
                            elif Az_ in C.original_files:
                                AX_ = D.get(w, SQL_UPDATE_TEMPLATE)
                                AY_ = I
                                AZ_ = I
                                try:
                                    import re

                                    BU = re.search(
                                        "(?i)update\\s+([0-9A-Za-z_\\.]+)\\s+set", AX_
                                    )
                                    if BU:
                                        AY_ = BU.group(1)
                                    BV = AX_.lower().find(" where")
                                    if BV != -1:
                                        AZ_ = AX_[BV:]
                                except E:
                                    AY_ = I
                                    AZ_ = I
                                if not AY_:
                                    AY_ = "object_query_1"
                                if not AZ_:
                                    AZ_ = " WHERE EAN = '{ean}' OR Towar_powiazany_z_SKU = '{ean}'"
                                Bv = AZ_.replace("{ean}", K_)
                                AC_ = f"UPDATE {AY_} SET {B3_} = ''" + Bv
                                cur.execute(AC_)
                                Aq_ += 1
                                if Aj(cur, A3, -1) >= 0:
                                    CANCEL_LABEL += cur.rowcount
                        if Aq_ > 0:
                            conn.commit()
                            if Aq_:
                                log_info_loc(
                                    "db_update_success_log",
                                    ean=K_,
                                    cols=AI.join([f"{B3_} = ..." for B3_ in []]),
                                )
                        cur.close()
                        conn.close()
                    except E as R:
                        AW_ = G(R)
                        if "cur" in locals():
                            try:
                                cur.close()
                            except:
                                pass
                        if "conn" in locals():
                            try:
                                conn.rollback()
                            except:
                                pass
                            try:
                                conn.close()
                            except:
                                pass
                        log_error(f"SQL update error for EAN {K_}: {R}")
                    INCOMPLETE_DATA_MSG = int((Ag.perf_counter() - Br) * 1000)
                result_data[P] = AW_
                result_data[d] = Aq_
                result_data[j] = CANCEL_LABEL
                result_data[S] = INCOMPLETE_DATA_MSG
            except E as exc:
                log_error_loc(
                    "processing_unexpected_error", error=exc
                )
                result_data[K] = set(range(len(C.slots)))
                result_data[Y] = "Operacja przerwana z powodu błędu."
                result_data[P] = G(exc)
            result_data["ean"] = K_
            result_data[A2] = BY_

        thread = threading.Thread(target=heavy_work)
        thread.daemon = True
        thread.start()

        def check_thread():
            if thread.is_alive():
                C.after(100, check_thread)
            else:
                finalize()

        C.after(100, check_thread)

        def finalize():
            A = WARNING_LABEL
            for widget in [
                C.combo_name,
                C.combo_type,
                C.combo_model,
                C.combo_color1,
                C.combo_color2,
                C.combo_color3,
                C.combo_extra,
                C.entry_ean,
            ]:
                try:
                    widget.configure(state=X)
                except:
                    pass
            C.btn_submit.configure(state=X)
            C.btn_open.configure(state=X)
            C.is_processing = h
            err_set = result_data.get(K, set()) or set()
            add_set = result_data.get(p, set())
            del_set = result_data.get(s, set())
            inter_set = result_data.get(t, set())
            for F_ in err_set:
                C._mark_slot(F_, Ab)
            for F_ in inter_set:
                if F_ not in err_set:
                    C._mark_slot(F_, A4)
            for F_ in add_set - inter_set:
                if F_ not in err_set:
                    C._mark_slot(F_, A4)
            for F_ in del_set - inter_set:
                if F_ not in err_set:
                    C._mark_slot(F_, "gray")
            for F_, d_ in A0(C.slots):
                if F_ in add_set or F_ in del_set or F_ in err_set:
                    continue
                if d_[f]:
                    C._mark_slot(F_, A4)
                else:
                    C._mark_slot(F_, I)
            C.pending_additions = result_data.get(n, {})
            C.pending_deletions = result_data.get(o, {})
            Y_ = result_data.get(Y, B)
            A__ = result_data.get(k, Ay)
            AW_msg = result_data.get(P, B)
            K_val = result_data.get("ean", K_)
            if not err_set and not A__ and not Y_ and not AW_msg:
                C._load_existing_files()
            if err_set:
                O.showwarning(
                    A,
                    OPERATION_ERRORS_MSG.format(backup=AN),
                )
            elif Y_:
                O.showerror(
                    FTP_ERROR_LABEL,
                    FTP_SEND_FAILED_MSG.format(reason=Y_),
                )
            elif A__:
                O.showwarning(
                    A,
                    FTP_SKIPPED_NO_EAN_MSG,
                )
            elif result_data[P]:
                O.showerror(
                    SQL_ERROR_LABEL,
                    SQL_UPDATE_FAILED_MSG.format(reason=result_data[P]),
                )
            else:
                O.showinfo(SAVED_LABEL, UPDATE_SUCCESS_MSG.format(ean=K_val))
            if not A__:
                status = "OK" if not Y_ else Y_
                log_info_loc(
                    "ftp_summary",
                    uploaded=result_data[Z],
                    deleted=result_data[b],
                    time=result_data[c],
                    status=status,
                )
            if D.get(u, J):
                if result_data[P]:
                    log_info_loc(
                        "sql_error", error=result_data[P], time=result_data[S]
                    )
                else:
                    log_info_loc(
                        "sql_summary",
                        queries=result_data[d],
                        rows=result_data[j],
                        time=result_data[S],
                    )
            if result_data.get(A2, Ay):
                log_info_loc(
                    "entry_updated_log",
                    ean=K_val,
                    name=AE_,
                    type=AF_,
                    model=AG_,
                    color1=AH_,
                    color2=p_,
                    color3=s_,
                    extras=b_,
                )
            else:
                log_info_loc(
                    "entry_added_log",
                    ean=K_val,
                    name=AE_,
                    type=AF_,
                    model=AG_,
                    color1=AH_,
                    color2=p_,
                    color3=s_,
                    extras=b_,
                )

    def _load_by_ean(A):
        E_ = NO_EAN_LABEL
        D_ = A.var_ean.get().strip()
        if not D_:
            O.showwarning(E_, ENTER_EAN_TO_LOAD_MSG)
            return
        if D_.upper() == q:
            O.showwarning(E_, CANNOT_SEARCH_NO_EAN_MSG)
            return
        if D_ in A.entries:
            C_ = A.entries[D_]
            G_ = C_.get(Ae, B) or B
            H_ = C_.get(Ad, B) or B
            I_ = C_.get(AZ, B) or B
            K_ = C_.get(AY, B) or B
            M_ = C_.get(AX, B) or B
            N_ = C_.get(AW, B) or B
            F_ = C_.get(d, B) or B
            A.suppress_scan = J
            try:
                A.var_name.set(G_)
                A._on_name_commit()
                A.var_type.set(H_)
                A._on_type_commit()
                A.var_model.set(I_)
                A.loading_by_ean = J
                A._on_model_commit()
                A.loading_by_ean = h
                A.var_color1.set(K_)
                A.var_color2.set(M_)
                A.var_color3.set(N_)
                A._on_color_commit()
                if F_.upper() == L:
                    A.var_extra.set(B)
                else:
                    A.var_extra.set(F_)
                A._on_extra_commit()
                A.var_ean.set(D_)
            finally:
                A.suppress_scan = h
            A._load_existing_files()
        else:
            A._load_existing_files()
            O.showinfo(NOT_FOUND_LABEL, NO_SAVED_DATA_FOR_EAN_MSG.format(ean=D_))

    def _open_current_folder(B):
        F_ = B.var_name.get().strip()
        G_ = B.var_type.get().strip()
        H_ = B.var_model.get().strip()
        I_ = B.var_color1.get().strip()
        K_ = B.var_color2.get().strip()
        M_ = B.var_color3.get().strip()
        N_ = B.var_extra.get().strip()
        if not (F_ and G_ and H_ and I_):
            O.showwarning(
                NO_DATA_MSG,
                FILL_REQUIRED_BEFORE_OPEN_MSG,
            )
            return
        C_ = A.path.join(l, F_.upper(), G_.upper(), H_.upper())
        D_ = [I_.upper()]
        if K_:
            D_.append(K_.upper())
        if M_:
            D_.append(M_.upper())
        Q_ = g.join(D_)
        R_ = N_.strip().replace(a, g).upper() if N_ else L
        C_ = A.path.join(C_, Q_, R_)
        A.makedirs(C_, exist_ok=J)
        try:
            if A.name == "nt":
                A.startfile(C_)
            else:
                BH.run(["xdg-open", C_], check=h)
        except E as P_:
            O.showerror(AK, FOLDER_OPEN_FAILED_MSG.format(error=P_))
            log_error_loc("folder_open_error", path=C_, error=P_)

    def _open_settings(A):
        a = CHANGE_DATA_ADMIN_LABEL
        Y = "*"
        i_ = "readonly"
        A5_ = RUN_AS_ADMIN_MSG
        A6_ = NO_PERMISSIONS_LABEL
        Ag_ = a
        A7_ = DATABASE_LABEL
        A8_ = SERVER_LABEL
        A9_ = MSSQL_SERVER_LABEL
        AA_ = TEST_BUTTON_LABEL
        AC_ = CONNECTED_LABEL
        j_ = PASSWORD_LABEL
        k_ = USER_LABEL
        f_ = MYSQL_LABEL
        Y_ = "write"
        d_ = i_
        a_ = F.Toplevel(A)
        a_.title(SETTINGS_LABEL)
        a_.grab_set()
        Z = C.Notebook(a_)
        Z.pack(expand=J, fill=z, padx=5, pady=5)
        L = C.Frame(Z)
        Q = C.Frame(Z)
        S = C.Frame(Z)
        U = C.Frame(Z)
        Z.add(L, text=IMAGES_TAB_LABEL)
        Z.add(Q, text=FTP_TAB_LABEL)
        Z.add(S, text=SQL_TAB_LABEL)
        Z.add(U, text=LANGUAGE_TAB_LABEL)
        C.Label(L, text=IMAGE_SETTINGS_LABEL).grid(
            row=0, column=0, columnspan=4, padx=5, pady=5, sticky=T
        )
        Ah = C.Checkbutton(L, text=B, variable=A.opt_resize)
        Ah.grid(row=1, column=0, padx=5, sticky=T)
        C.Label(L, text=RESIZE_LABEL).grid(row=1, column=1, sticky=T)
        l_ = C.Entry(L, textvariable=A.resize_max_dim, width=5)
        l_.grid(row=1, column=2)
        C.Label(L, text=PX_MAX_LABEL).grid(row=1, column=3, sticky=T)
        Ai = C.Checkbutton(L, text=B, variable=A.opt_compress)
        Ai.grid(row=2, column=0, padx=5, sticky=T)
        C.Label(L, text=COMPRESS_LABEL).grid(row=2, column=1, sticky=T)
        n = C.Spinbox(L, from_=10, to=100, textvariable=A.compress_quality, width=5)
        n.grid(row=2, column=2, sticky=T)
        C.Label(L, text="%").grid(row=2, column=3, sticky=T)
        Aj = C.Checkbutton(L, text=B, variable=A.opt_maxsize)
        Aj.grid(row=3, column=0, padx=5, sticky=T)
        C.Label(L, text=LIMIT_SIZE_LABEL).grid(row=3, column=1, sticky=T)
        o = C.Spinbox(
            L, from_=100, to=10000, increment=100, textvariable=A.max_file_kb, width=6
        )
        o.grid(row=3, column=2, sticky=T)
        C.Label(L, text="KB").grid(row=3, column=3, sticky=T)
        Ak = C.Checkbutton(L, text=B, variable=A.opt_convert_tif)
        Ak.grid(row=4, column=0, padx=5, sticky=T)
        C.Label(L, text=CONVERT_TIF_LABEL).grid(row=4, column=1, sticky=T)
        q = C.Combobox(
            L,
            textvariable=A.tif_target_format,
            values=[At, "JPG", "BMP", "GIF"],
            state=d_,
            width=5,
        )
        q.grid(row=4, column=2, sticky=T)
        C.Label(U, text=LANGUAGE_LABEL).grid(row=0, column=0, sticky=R, padx=5, pady=2)
        lang_var = F.StringVar(value=LANG_PREF)
        lang_combo = C.Combobox(
            U,
            textvariable=lang_var,
            values=["auto", "pl", "ua", "en"],
            state="readonly",
            width=10,
        )
        lang_combo.grid(row=0, column=1, padx=5, pady=2, sticky=T)
        lang_combo.configure(postcommand=lambda c=lang_combo: A._style_combobox_list(c))
        C.Label(U, text=LOC_PATH_LABEL).grid(row=1, column=0, sticky=R, padx=5, pady=2)
        loc_var = F.StringVar(value=D.get("loc_path", LC))
        loc_entry = C.Entry(U, textvariable=loc_var, width=40, state=i_)
        loc_entry.grid(row=1, column=1, padx=5, pady=2, sticky=T)

        def browse_loc():
            A_ = BT.askdirectory(title=LOC_PATH_LABEL)
            if A_:
                loc_var.set(A_)

        browse_btn = C.Button(U, text=CHOOSE_LABEL, command=browse_loc, state=V)
        browse_btn.grid(row=1, column=2, padx=5, pady=2)

        C.Label(U, text=LOC_URLS_LABEL).grid(row=2, column=0, sticky=R, padx=5, pady=2)
        url_text = BS.ScrolledText(U, width=80, height=5, state=V, wrap="none")
        url_text.grid(row=2, column=1, padx=5, pady=2, sticky=T)
        url_text.configure(state=X)
        url_text.insert(A_, "\n".join(D.get("loc_urls", LOC_URLS)))
        url_text.configure(state=V)
        update_btn = C.Button(U, text=UPDATE_LOC_LABEL, state=V)
        update_btn.grid(row=2, column=2, padx=5, pady=2)

        def update_loc():
            global LOC_URLS, LC
            urls = [u.strip() for u in url_text.get(A_, "end").splitlines() if u.strip()]
            LOC_URLS = urls
            LC = loc_var.get().strip()
            if download_localizations(J):
                O.showinfo(SETTINGS_LABEL, LOC_UPDATE_SUCCESS_MSG)
            else:
                O.showwarning(
                    SETTINGS_LABEL,
                    LANG.get(
                        "loc_download_failed",
                        "Localization files unavailable. Check repository access.",
                    ),
                )

        update_btn.configure(command=update_loc)

        lang_unlock_btn = C.Button(U, text=a)
        lang_unlock_btn.grid(row=3, column=0, sticky=R, padx=5, pady=5)

        def unlock_lang():
            if is_admin():
                loc_entry.configure(state=X)
                browse_btn.configure(state=X)
                url_text.configure(state=X)
                update_btn.configure(state=X)
            else:
                O.showwarning(A6_, A5_)

        lang_unlock_btn.configure(command=unlock_lang)

        def Am(*B):
            l_.configure(state=X if A.opt_resize.get() else V)

        def An(*B):
            n.configure(state=X if A.opt_compress.get() else V)

        def Ao(*B):
            o.configure(state=X if A.opt_maxsize.get() else V)

        Ap = A.opt_resize.trace_add(Y_, lambda *A_: Am())
        Aq = A.opt_compress.trace_add(Y_, lambda *A_: An())
        Ar = A.opt_maxsize.trace_add(Y_, lambda *B: Ao())
        get_file_lock_user = A.opt_convert_tif.trace_add(
            Y_, lambda *B: q.configure(state=d_ if A.opt_convert_tif.get() else V)
        )
        l_.configure(state=X if A.opt_resize.get() else V)
        n.configure(state=X if A.opt_compress.get() else V)
        o.configure(state=X if A.opt_maxsize.get() else V)
        q.configure(state=d_ if A.opt_convert_tif.get() else V)
        C.Label(Q, text=FTP_SERVER_LABEL).grid(row=0, column=0, sticky=R, padx=5, pady=2)
        s = F.StringVar(value=D[H][v])
        AD_ = C.Entry(Q, textvariable=s, width=30)
        AD_.grid(row=0, column=1, padx=5, pady=2)
        C.Label(Q, text=PORT_LABEL).grid(row=1, column=0, sticky=R, padx=5, pady=2)
        t = F.IntVar(value=D[H][r])
        AE_ = C.Entry(Q, textvariable=t, width=6)
        AE_.grid(row=1, column=1, sticky=T, padx=5, pady=2)
        C.Label(Q, text=k_).grid(row=2, column=0, sticky=R, padx=5, pady=2)
        x_ = F.StringVar(value=D[H][N])
        AF_ = C.Entry(Q, textvariable=x_, width=30)
        AF_.grid(row=2, column=1, padx=5, pady=2)
        C.Label(Q, text=j_).grid(row=3, column=0, sticky=R, padx=5, pady=2)
        y_ = F.StringVar(value=D[H][M])
        AG_ = C.Entry(Q, textvariable=y_, show=Y, width=30)
        AG_.grid(row=3, column=1, padx=5, pady=2)
        C.Label(Q, text=FTP_PATH_LABEL).grid(
            row=4, column=0, sticky=R, padx=5, pady=2
        )
        g_ = F.StringVar(value=D[H][m])
        AH_ = C.Entry(Q, textvariable=g_, width=30)
        AH_.grid(row=4, column=1, padx=5, pady=2)
        AI_ = C.Button(Q, text=a)
        AI_.grid(row=5, column=0, sticky=R, padx=5, pady=5)
        C.Label(Q, text=FTP_TEST_LABEL).grid(
            row=6, column=0, sticky=R, padx=5, pady=5
        )
        AJ_ = F.StringVar(value=B)
        Aw = C.Entry(Q, textvariable=AJ_, width=50, state=d_)
        Aw.grid(row=6, column=1, padx=5, pady=5, sticky=T)

        def Ax():
            A_ = B
            try:
                C_ = AB.FTP()
                C_.connect(s.get(), t.get(), timeout=10)
                C_.login(x_.get(), y_.get())
                C_.set_pasv(J)
                if g_.get():
                    C_.cwd(g_.get())
            except AB.error_perm as F_:
                D_ = G(F_)
                if "530" in D_ or LOGIN_INCORRECT_MSG in D_:
                    A_ = LOGIN_DATA_ERROR_MSG
                elif As in D_ or NO_SUCH_FILE_MSG in D_:
                    A_ = PATH_NOT_FOUND_MSG
                else:
                    A_ = FTP_GENERIC_ERROR_MSG.format(error=D_)
            except (BK.gaierror, CONNECTION_REFUSED_ERROR, TIMEOUT_ERROR, Au) as F_:
                A_ = NETWORK_ERROR_MSG
            except E as F_:
                A_ = OTHER_ERROR_MSG.format(error=F_)
            else:
                A_ = AC_
                try:
                    C_.quit()
                except E:
                    pass
            AJ_.set(A_)

        Ay = C.Button(Q, text=AA_, command=Ax)
        Ay.grid(row=6, column=1, padx=5, pady=5, sticky=R)
        C.Label(Q, text=FTP_UPDATE_LABEL).grid(
            row=7, column=0, sticky=R, padx=5, pady=2
        )
        ftp_update_var = F.BooleanVar(value=D.get(ft, J))
        ftp_update_cb = C.Checkbutton(Q, variable=ftp_update_var)
        ftp_update_cb.grid(row=7, column=1, sticky=T, padx=5, pady=2)
        C.Label(S, text=DB_TYPE_LABEL).grid(
            row=0, column=0, sticky=R, padx=5, pady=2
        )
        A0 = F.StringVar(value=f_ if D.get(p, K).lower() == K else A9_)
        A1 = C.Combobox(S, textvariable=A0, values=[A9_, f_], state=d_, width=20)
        A1.grid(row=0, column=1, padx=5, pady=2, sticky=T)
        U = C.Frame(S)
        W = C.Frame(S)
        C.Label(U, text=A8_).grid(row=0, column=0, sticky=R, padx=5, pady=2)
        AK = F.StringVar(value=D[P][c])
        ensure_package = C.Entry(U, textvariable=AK, width=30)
        ensure_package.grid(row=0, column=1, padx=5, pady=2)
        C.Label(U, text=A7_).grid(row=1, column=0, sticky=R, padx=5, pady=2)
        AM = F.StringVar(value=D[P][b])
        AN = C.Entry(U, textvariable=AM, width=30)
        AN.grid(row=1, column=1, padx=5, pady=2)
        C.Label(U, text=k_).grid(row=2, column=0, sticky=R, padx=5, pady=2)
        AO = F.StringVar(value=D[P][N])
        AQ = C.Entry(U, textvariable=AO, width=30)
        AQ.grid(row=2, column=1, padx=5, pady=2)
        C.Label(U, text=j_).grid(row=3, column=0, sticky=R, padx=5, pady=2)
        AR = F.StringVar(value=D[P][M])
        AS = C.Entry(U, textvariable=AR, show=Y, width=30)
        AS.grid(row=3, column=1, padx=5, pady=2)
        U.grid(row=1, column=0, columnspan=2, sticky=T, padx=5, pady=2)
        C.Label(W, text=A8_).grid(row=0, column=0, sticky=R, padx=5, pady=2)
        AT = F.StringVar(value=D[K][c])
        AU = C.Entry(W, textvariable=AT, width=30)
        AU.grid(row=0, column=1, padx=5, pady=2)
        C.Label(W, text=A7_).grid(row=1, column=0, sticky=R, padx=5, pady=2)
        AV = F.StringVar(value=D[K][b])
        AW = C.Entry(W, textvariable=AV, width=30)
        AW.grid(row=1, column=1, padx=5, pady=2)
        C.Label(W, text=k_).grid(row=2, column=0, sticky=R, padx=5, pady=2)
        AX = F.StringVar(value=D[K][N])
        AY = C.Entry(W, textvariable=AX, width=30)
        AY.grid(row=2, column=1, padx=5, pady=2)
        C.Label(W, text=j_).grid(row=3, column=0, sticky=R, padx=5, pady=2)
        AZ = F.StringVar(value=D[K][M])
        Aa = C.Entry(W, textvariable=AZ, show=Y, width=30)
        Aa.grid(row=3, column=1, padx=5, pady=2)
        W.grid(row=1, column=0, columnspan=2, sticky=T, padx=5, pady=2)
        if D.get(p, K).lower() == K:
            U.grid_remove()
        else:
            W.grid_remove()

        def Az(event=I):
            if A0.get() == f_:
                U.grid_remove()
                W.grid()
            else:
                W.grid_remove()
                U.grid()

        A1.bind(A2, Az)
        C.Label(S, text=SQL_UPDATE_LABEL).grid(
            row=2, column=0, sticky=R, padx=5, pady=2
        )
        Ab = F.BooleanVar(value=D.get(u, J))
        Ac = C.Checkbutton(S, variable=Ab)
        Ac.grid(row=2, column=1, sticky=T, padx=5, pady=2)
        C.Label(S, text=SQL_QUERY_LABEL).grid(
            row=3, column=0, sticky="ne", padx=5, pady=2
        )
        h_ = F.Text(S, width=80, height=3)
        h_.insert(A_, D.get(w, SQL_UPDATE_TEMPLATE))
        h_.grid(row=3, column=1, padx=5, pady=2, sticky=T)
        C.Label(S, text=SQL_TEST_LABEL).grid(
            row=4, column=0, sticky=R, padx=5, pady=5
        )
        A3_ = F.StringVar(value=B)
        MISSING_FIELDS_MSG = C.Entry(S, textvariable=A3_, width=50, state=d_)
        MISSING_FIELDS_MSG.grid(row=4, column=1, padx=5, pady=5, sticky=T)

        def INCOMPLETE_DATA_MSG():
            try:
                A_ = connect_db()
                try:
                    B_ = A_.cursor()
                    try:
                        B_.execute("SELECT 1")
                    except E:
                        pass
                    finally:
                        try:
                            B_.close()
                        except E:
                            pass
                finally:
                    try:
                        A_.close()
                    except E:
                        pass
                A3_.set(AC_)
            except E as C_:
                A3_.set(f"Błąd: {C_}")

        EDIT_LISTS_LABEL = C.Button(S, text=AA_, command=INCOMPLETE_DATA_MSG)
        EDIT_LISTS_LABEL.grid(row=4, column=1, padx=5, pady=5, sticky=R)

        def Ad(state):
            A_ = state
            AD_.configure(state=A_)
            AE_.configure(state=A_)
            AF_.configure(state=A_)
            AG_.configure(state=A_)
            AH_.configure(state=A_)

        def Ae(state_text, editor=Al):
            B_ = state_text
            A__ = B_
            C_ = X if B_ == X else V
            if D.get(p, K).lower() == K:
                AU.configure(state=A__)
                AW.configure(state=A__)
                AY.configure(state=A__)
                Aa.configure(state=A__)
            else:
                ensure_package.configure(state=A__)
                AN.configure(state=A__)
                AQ.configure(state=A__)
                AS.configure(state=A__)
            A1.configure(state=i_ if editor else A__)
            h_.configure(state=C_)
            Ac.configure(state=X)

        Ad(i_)
        Ae(i_)

        def LIGHT_GREEN():
            if is_admin():
                Ad(X)
                log_info_loc("settings_ftp_unlocked")
            else:
                O.showwarning(A6_, A5_)

        def NO_DATA_MSG():
            if is_admin():
                Ae(X)
                log_info_loc("settings_sql_unlocked")
            else:
                O.showwarning(A6_, A5_)

        AI_.configure(command=LIGHT_GREEN)
        BC_ = C.Button(S, text=Ag_, command=NO_DATA_MSG)
        BC_.grid(row=5, column=1, sticky=T, padx=5, pady=5)
        A4 = C.Frame(a_)
        A4.pack(pady=5)

        def BD_():
            global LC, LOC_URLS, LANG_PREF
            D[H][v] = s.get().strip()
            try:
                D[H][r] = int(t.get())
            except:
                D[H][r] = 21
            D[H][N] = x_.get().strip()
            D[H][M] = y_.get()
            D[H][m] = g_.get().strip()
            D[ft] = bool(ftp_update_var.get())
            D[P][c] = AK.get().strip()
            D[P][b] = AM.get().strip()
            D[P][N] = AO.get().strip()
            D[P][M] = AR.get()
            D[K][c] = AT.get().strip()
            D[K][b] = AV.get().strip()
            D[K][N] = AX.get().strip()
            D[K][M] = AZ.get()
            D[p] = K if A0.get() == f_ else "mssql"
            D[w] = h_.get(A_, "end").strip()
            D[u] = bool(Ab.get())
            D["loc_path"] = loc_var.get().strip()
            D["loc_urls"] = [u.strip() for u in url_text.get(A_, "end").splitlines() if u.strip()]
            LC = D["loc_path"]
            LOC_URLS = D["loc_urls"]
            save_language_pref(lang_var.get().strip())
            LANG_PREF = lang_var.get().strip()
            save_config(D)
            log_info_loc("settings_saved")
            Af()

        C.Button(A4, text=SAVE_LABEL, command=BD_).grid(row=0, column=0, padx=5)

        def Af():
            A.opt_resize.trace_remove(Y_, Ap)
            A.opt_compress.trace_remove(Y_, Aq)
            A.opt_maxsize.trace_remove(Y_, Ar)
            A.opt_convert_tif.trace_remove(Y_, get_file_lock_user)
            a_.destroy()

        C.Button(A4, text=CANCEL_LABEL, command=Af).grid(row=0, column=1, padx=5)
        a_.protocol("WM_DELETE_WINDOW", Af)
        Z.select(0)

    def _change_language(A):
        B = BI.askstring(SETTINGS_LABEL, LANGUAGE_PROMPT)
        if B:
            try:
                save_language_pref(B.lower())
            except E:
                O.showerror(AK, Ac)
            else:
                O.showinfo(SETTINGS_LABEL, RESTART_TO_APPLY_LABEL)

    def _style_combobox_list(L, combobox):
        A_ = combobox
        try:
            G_ = A_.tk.call("ttk::combobox::PopdownWindow", A_._w)
            H_ = G_ + ".f.l"
            B_ = A_.nametowidget(H_)
        except E:
            return
        D_ = Aj(A_, "existing_count", I)
        if D_ is I:
            return
        F_ = A_.cget(S)
        J_ = Q(F_) if F_ else 0
        K_ = B_.cget("background")
        for C_ in Ax(J_):
            if C_ < D_:
                B_.itemconfig(C_, background=LIGHT_GREEN)
            else:
                B_.itemconfig(C_, background=K_)

    def _mark_slot(D, idx, color):
        B_ = color
        E_ = {AR: "#0000ff", A4: "#00ff00", "gray": "#808080", Ab: "#ff0000"}
        C_ = E_.get(B_, "#000000")
        slot = D.slots[idx]
        slot[B0] = B_
        A_ = slot.get(AS)
        if A_:
            if B_ is I:
                A_.configure(
                    highlightthickness=0, highlightbackground=A8, highlightcolor=A8
                )
            else:
                A_.configure(
                    highlightbackground=C_, highlightcolor=C_, highlightthickness=2
                )

    def _add_tooltip(C, widget, text):
        B_ = widget
        A_ = I

        def D_(event):
            B__ = event
            nonlocal A_
            A_ = F.Toplevel(C)
            A_.wm_overrideredirect(J)
            A_.wm_geometry(f"+{B__.x_root+10}+{B__.y_root+10}")
            D__ = F.Label(
                A_,
                text=text,
                background="yellow",
                relief="solid",
                borderwidth=1,
                padx=5,
                pady=3,
            )
            D__.pack()

        def E__(event):
            nonlocal A_
            if A_:
                A_.destroy()
                A_ = I

        B_.bind("<Enter>", D_)
        B_.bind("<Leave>", E__)

    def _on_drag_init(A, event, idx):
        if A.is_processing:
            return
        B_ = A.slots[idx][f]
        if not B_:
            return
        A.dragging_idx = idx
        return "copy", BJ, (B_,)

    def _on_drag_end(A, event):
        A.dragging_idx = I

    def _ui_log(A, msg=AQ, clear=Ay):
        try:
            if clear:
                A.ui_log.configure(state=Az)
                A.ui_log.delete(A_, F.END)
                A.ui_log.configure(state=Ak)
                return
            if not msg:
                return
            A.ui_log.configure(state=Az)
            A.ui_log.insert(F.END, f"{msg}\n")
            A.ui_log.see(F.END)
            A.ui_log.configure(state=Ak)
        except E:
            pass


if __name__ == "__main__":
    A1 = App()
    if not LOC_DL_OK:
        O.showwarning(SETTINGS_LABEL, LANG_EN.get("loc_download_failed", "Localization files unavailable. Check repository access."))
    for BQ in (
        A1.combo_name,
        A1.combo_type,
        A1.combo_model,
        A1.combo_color1,
        A1.combo_color2,
        A1.combo_color3,
        A1.combo_extra,
    ):
        BQ.configure(postcommand=lambda c=BQ: A1._style_combobox_list(c))
    A1.mainloop()
